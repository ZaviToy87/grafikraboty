# -*- coding: utf-8 -*-
"""
web_converter.py — Конвертер ценников API

Функции:
- Загрузка прайс-листа
- Анализ цен
- Генерация ценников
- Печать
"""
from flask import Blueprint, request, jsonify, session, send_file, render_template, redirect
from datetime import datetime
from web_config import logger, get_db_connection, UPLOADS_DIR
import os
import json

converter_bp = Blueprint('converter', __name__)

# Папка для загрузок конвертера
CONVERTER_UPLOADS_DIR = os.path.join(UPLOADS_DIR, 'converter')
CONVERTER_OUTPUT_DIR = os.path.join(UPLOADS_DIR, 'price_tags')  # Папка для готовых ценников
os.makedirs(CONVERTER_UPLOADS_DIR, exist_ok=True)
os.makedirs(CONVERTER_OUTPUT_DIR, exist_ok=True)


@converter_bp.route('/converter')
def converter_page():
    """Страница конвертера ценников"""
    if 'user_id' not in session:
        return redirect('/login')
    return render_template('price_converter.html')


@converter_bp.route('/converter/files', methods=['GET'])
def get_converter_files():
    """Получить список файлов конвертера"""
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401
    
    files = []
    
    # Готовые файлы (ценники)
    if os.path.exists(CONVERTER_OUTPUT_DIR):
        for f in os.listdir(CONVERTER_OUTPUT_DIR):
            if f.endswith('.xlsx'):
                filepath = os.path.join(CONVERTER_OUTPUT_DIR, f)
                meta_path = os.path.join(CONVERTER_OUTPUT_DIR, f.replace('.xlsx', '.json'))
                
                # Пытаемся загрузить метаданные
                meta = {}
                if os.path.exists(meta_path):
                    try:
                        with open(meta_path, 'r', encoding='utf-8') as mf:
                            meta = json.load(mf)
                    except:
                        pass
                
                files.append({
                    'filename': f,
                    'path': filepath,
                    'size': os.path.getsize(filepath),
                    'created_at': meta.get('created_at', datetime.fromtimestamp(os.path.getctime(filepath)).isoformat()),
                    'products_count': meta.get('products_count', 0),
                    'settings': meta.get('settings', {}),
                    'type': 'output'  # Готовый файл
                })
    
    # Загруженные файлы (прайсы)
    if os.path.exists(CONVERTER_UPLOADS_DIR):
        for f in os.listdir(CONVERTER_UPLOADS_DIR):
            if f.endswith(('.xlsx', '.xls', '.csv')):
                filepath = os.path.join(CONVERTER_UPLOADS_DIR, f)
                files.append({
                    'filename': f,
                    'path': filepath,
                    'size': os.path.getsize(filepath),
                    'created_at': datetime.fromtimestamp(os.path.getctime(filepath)).isoformat(),
                    'type': 'input'  # Исходный файл
                })
    
    # Сортируем по дате (новые сверху)
    files.sort(key=lambda x: x['created_at'], reverse=True)
    
    return jsonify({'status': 'success', 'files': files})


@converter_bp.route('/converter/download/<filename>', methods=['GET'])
def download_price_tag(filename):
    """Скачать готовый файл с ценниками"""
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401
    
    # Безопасное имя файла
    safe_filename = os.path.basename(filename)
    filepath = os.path.join(CONVERTER_OUTPUT_DIR, safe_filename)
    
    if not os.path.exists(filepath):
        return jsonify({'status': 'error', 'message': 'Файл не найден'}), 404
    
    return send_file(
        filepath,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=False,  # Открывать в браузере, а не скачивать
        download_name=safe_filename
    )


@converter_bp.route('/converter/view/<filename>', methods=['GET'])
def view_price_tag(filename):
    """Открыть файл для просмотра (скачивание)"""
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401
    
    safe_filename = os.path.basename(filename)
    filepath = os.path.join(CONVERTER_OUTPUT_DIR, safe_filename)
    
    if not os.path.exists(filepath):
        return jsonify({'status': 'error', 'message': 'Файл не найден'}), 404
    
    return send_file(
        filepath,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,  # Скачивать файл
        download_name=safe_filename
    )


@converter_bp.route('/converter/upload', methods=['POST'])
def upload_price_file():
    """Загрузить прайс-лист"""
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401
    
    if 'file' not in request.files:
        return jsonify({'status': 'error', 'message': 'Файл не загружен'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'status': 'error', 'message': 'Файл не выбран'}), 400
    
    # Сохраняем файл
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    safe_name = f"{timestamp}_{file.filename}"
    filepath = os.path.join(CONVERTER_UPLOADS_DIR, safe_name)
    file.save(filepath)
    
    logger.info(f"Price file uploaded: {safe_name}")
    
    return jsonify({
        'status': 'success',
        'message': 'Файл загружен',
        'filename': safe_name
    })


@converter_bp.route('/converter/analyze', methods=['POST'])
def analyze_prices():
    """Анализировать цены из файла"""
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401

    data = request.json or {}
    filename = data.get('filename')

    if not filename:
        return jsonify({'status': 'error', 'message': 'Filename required'}), 400

    filepath = os.path.join(CONVERTER_UPLOADS_DIR, filename)

    if not os.path.exists(filepath):
        return jsonify({'status': 'error', 'message': 'File not found'}), 404

    try:
        # Определяем тип файла
        ext = os.path.splitext(filename)[1].lower()

        if ext in ['.xlsx', '.xls']:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(filepath, read_only=True)
                ws = wb.active

                products = []
                
                # Читаем все строки и ищем данные
                # Формат: колонка A (0) - название, колонка H (7) - новая цена
                for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                    # Пропускаем первые 4 строки (заголовки)
                    if i <= 4:
                        continue
                    
                    # Получаем название из колонки A
                    name = row[0] if len(row) > 0 else None
                    
                    # Пропускаем пустые строки
                    if not name or str(name).strip() == '':
                        continue
                    
                    # Получаем новую розничную цену — берём САМУЮ ПРАВУЮ колонку с ценой
                    # Ищем последнюю числовую колонку в строке (новая цена всегда правее)
                    price = None
                    for col_idx in range(len(row) - 1, 0, -1):  # Идём справа налево
                        val = row[col_idx]
                        if val is not None and val != '':
                            try:
                                price = float(val)
                                break  # Нашли цену — выходим
                            except (ValueError, TypeError):
                                continue  # Не число — пропускаем

                    if price is None or price == 0:
                        continue  # Нет цены — пропускаем строку
                    
                    products.append({
                        'row': i,
                        'name': str(name).strip(),
                        'price': float(price),
                        'unit': 'шт'
                    })

                wb.close()
                
                if len(products) == 0:
                    return jsonify({
                        'status': 'error', 
                        'message': 'Нет валидных данных. Ожидаемый формат: колонка A - наименование, последняя колонка с числом - новая розничная цена. Строки данных начинаются после заголовков.'
                    }), 400

            except Exception as e:
                logger.exception(f"Excel parse error: {e}")
                return jsonify({'status': 'error', 'message': f'Ошибка чтения Excel: {str(e)}'}), 500

        elif ext == '.csv':
            import csv
            products = []
            with open(filepath, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                for i, row in enumerate(reader, 1):
                    if i <= 4:  # Пропускаем заголовки
                        continue
                    if row and len(row) > 0 and row[0]:
                        # Ищем новую цену — самая правая колонка с числом
                        price = None
                        for col_idx in range(len(row) - 1, 0, -1):
                            try:
                                price = float(row[col_idx])
                                break
                            except (ValueError, TypeError):
                                continue
                        
                        if price is None or price == 0:
                            continue
                        
                        products.append({
                            'row': i,
                            'name': row[0],
                            'price': price,
                            'unit': 'шт'
                        })
        else:
            return jsonify({'status': 'error', 'message': 'Неподдерживаемый формат файла'}), 400

        return jsonify({
            'status': 'success',
            'products': products,
            'count': len(products)
        })

    except Exception as e:
        logger.exception(f"Analyze error: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


@converter_bp.route('/converter/generate', methods=['POST'])
def generate_price_tags():
    """Сгенерировать ценники в Excel (как в рабочей версии КонвертаторЦенников 2.0)"""
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401

    data = request.json or {}
    products = data.get('products', [])
    settings = data.get('settings', {})
    markup = float(data.get('markup', 0))  # Наценка %

    if not products:
        return jsonify({'status': 'error', 'message': 'Нет товаров'}), 400

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl import Workbook
        from openpyxl.worksheet.worksheet import Worksheet
        from math import ceil
        from datetime import datetime

        # Создаём новую книгу
        wb = Workbook()
        ws = wb.active
        ws.title = "Ценники"

        # Настройки из UI
        n_cols = int(settings.get('cols', 3))  # Количество ценников в ряду
        org_name = settings.get('org_name', 'ООО "КАКИЕ ЛЮДИ"')
        org_size = int(settings.get('org_size', 8))
        org_bold = settings.get('org_bold', True)
        name_size = int(settings.get('name_size', 10))
        name_bold = settings.get('name_bold', False)
        price_base = int(settings.get('price_base', 18))
        price_bold = settings.get('price_bold', True)
        date_size = int(settings.get('date_size', 11))
        date_bold = settings.get('date_bold', False)
        font_name = settings.get('font_name', 'Calibri')

        # Параметры как в рабочей версии
        TOTAL_COL_WIDTH = 35  # суммарная ширина ценника
        NAME_COL_WIDTH = 35   # ширина колонки наименования (увеличено с 26)
        PRICE_COL_WIDTH = 12  # ширина колонки цены (увеличено с 9)
        LINE_HEIGHT_FACTOR = 1.35  # Увеличено с 1.25 для большего пространства
        NAME_PADDING = 16  # Увеличено с 8 для отступа сверху/снизу
        MAX_NAME_LINES = 5  # Максимум строк для наименования (увеличено с 4)

        # Текущая дата
        date_str = datetime.now().strftime('%d.%m.%Y')

        # Подготовка данных
        items = [(p.get('name', ''), float(p.get('price', 0)), p.get('unit', 'шт')) for p in products]
        total = len(items)

        # Блок ценника: 2 колонки x 4 строки
        block_w = 2
        block_h = 4

        # Устанавливаем ширины колонок
        for col_block in range(n_cols):
            c_name = col_block * block_w + 1
            c_price = col_block * block_w + 2
            ws.column_dimensions[get_column_letter(c_name)].width = NAME_COL_WIDTH
            ws.column_dimensions[get_column_letter(c_price)].width = PRICE_COL_WIDTH

        # Границы
        thin = Side(style="thin", color="000000")
        medium = Side(style="medium", color="000000")

        # Выравнивания
        align_org = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_name = Alignment(horizontal="left", vertical="top", wrap_text=True)
        align_price = Alignment(horizontal="center", vertical="center", wrap_text=False, shrink_to_fit=True)
        align_date = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Функция переноса текста по ширине (улучшенная)
        def wrap_text_by_width(text: str, max_chars: int) -> str:
            """
            Переносит текст на строки по максимальной ширине
            Учитывает что разные символы имеют разную ширину
            """
            text = str(text).strip()
            if not text:
                return ""

            # Если текст короткий — возвращаем как есть
            if len(text) <= max_chars:
                return text

            words = text.split()
            if not words:
                return text

            lines = []
            current_line = []
            current_length = 0

            for word in words:
                word_len = len(word)

                # Если слово само длиннее max_chars — разбиваем его
                if word_len > max_chars:
                    # Добавляем текущую строку если есть
                    if current_line:
                        lines.append(" ".join(current_line))
                        current_line = []
                        current_length = 0

                    # Разбиваем длинное слово по символам
                    for i in range(0, len(word), max_chars):
                        chunk = word[i:i + max_chars]
                        if i + max_chars < len(word):
                            # Это не последний кусок
                            lines.append(chunk)
                        else:
                            # Последний кусок
                            current_line = [chunk]
                            current_length = len(chunk)
                    continue

                # Проверяем влезет ли слово в текущую строку
                if current_length == 0:
                    # Первое слово в строке
                    current_line = [word]
                    current_length = word_len
                elif current_length + 1 + word_len <= max_chars:
                    # Слово влезает с пробелом
                    current_line.append(word)
                    current_length += 1 + word_len
                else:
                    # Слово не влезает — заканчиваем строку
                    lines.append(" ".join(current_line))
                    current_line = [word]
                    current_length = word_len

            # Добавляем последнюю строку
            if current_line:
                lines.append(" ".join(current_line))

            return "\n".join(lines)

        # Перенос наименований и расчёт высоты строк
        name_max_chars = NAME_COL_WIDTH + PRICE_COL_WIDTH  # имя занимает всю ширину блока
        n_block_rows = ceil(total / n_cols) if total > 0 else 0
        max_lines_per_row = [1] * n_block_rows
        wrapped_names: list = [""] * total

        # Первый проход: считаем строки для каждого товара
        for i, (name, _price, _unit) in enumerate(items):
            wrapped = wrap_text_by_width(name, name_max_chars)
            lines_count = max(1, wrapped.count("\n") + 1)

            # Ограничиваем максимум строк до MAX_NAME_LINES
            if lines_count > MAX_NAME_LINES:
                # Обрезаем текст и добавляем многоточие
                lines = wrapped.split("\n")[:MAX_NAME_LINES]
                last_line = lines[-1]
                if len(last_line) > name_max_chars - 3:
                    lines[-1] = last_line[:name_max_chars - 3] + "..."
                wrapped = "\n".join(lines)
                lines_count = MAX_NAME_LINES

            wrapped_names[i] = wrapped
            rb = i // n_cols
            if rb < len(max_lines_per_row) and lines_count > max_lines_per_row[rb]:
                max_lines_per_row[rb] = lines_count

        # Второй проход: находим ГЛОБАЛЬНЫЙ максимум для ВСЕГО документа
        # Чтобы все ценники имели одинаковую высоту
        global_max_lines = max(max_lines_per_row) if max_lines_per_row else 1
        print(f"📏 Глобальный максимум строк: {global_max_lines}")
        print(f"📏 По рядам: {max_lines_per_row}")

        # Устанавливаем ВСЕМ рядам одинаковую высоту по глобальному максимуму
        for rb in range(n_block_rows):
            max_lines_per_row[rb] = global_max_lines

        # Шрифты
        org_font = Font(name=font_name, size=org_size, bold=bool(org_bold))
        name_font = Font(name=font_name, size=name_size, bold=bool(name_bold))
        date_font = Font(name=font_name, size=date_size, bold=bool(date_bold))

        org_row_height = int(org_size * LINE_HEIGHT_FACTOR + 6)
        date_row_height = int(date_size * LINE_HEIGHT_FACTOR + 6)

        # Высоты строк — все одинаковые по глобальному максимуму
        for rb in range(n_block_rows):
            r0 = rb * block_h + 1
            ws.row_dimensions[r0].height = org_row_height

            # Высота для наименования: высота текста + отступы сверху и снизу
            lines_count = max_lines_per_row[rb]  # Теперь одинаковое для всех
            text_height = name_size * LINE_HEIGHT_FACTOR * lines_count
            total_padding = NAME_PADDING * 2  # Отступы сверху и снизу
            name_row_height = int(text_height + total_padding)

            # Минимальная высота
            min_height = int(name_size * LINE_HEIGHT_FACTOR + NAME_PADDING * 2)
            if name_row_height < min_height:
                name_row_height = min_height

            # Запас 10% для надёжности
            name_row_height = int(name_row_height * 1.1)

            ws.row_dimensions[r0 + 1].height = name_row_height

            # Цена и дата — фиксированная высота
            price_row_height = int(price_base * LINE_HEIGHT_FACTOR + 10)
            ws.row_dimensions[r0 + 2].height = price_row_height
            ws.row_dimensions[r0 + 3].height = date_row_height

        # Вывод блоков
        for i, (name, price, unit) in enumerate(items):
            rb = i // n_cols
            cb = i % n_cols
            r0 = rb * block_h + 1
            c0 = cb * block_w + 1

            # Границы блока: внешние medium, внутренние thin
            for rr in range(r0, r0 + block_h):
                for cc in range(c0, c0 + block_w):
                    left = medium if cc == c0 else thin
                    right = medium if cc == c0 + block_w - 1 else thin
                    top = medium if rr == r0 else thin
                    bottom = medium if rr == r0 + block_h - 1 else thin
                    ws.cell(row=rr, column=cc).border = Border(left=left, right=right, top=top, bottom=bottom)

            # r0: Организация (шапка) - объединено на 2 колонки
            ws.merge_cells(start_row=r0, start_column=c0, end_row=r0, end_column=c0 + 1)
            cell_org = ws.cell(row=r0, column=c0)
            cell_org.value = org_name
            cell_org.font = org_font
            cell_org.alignment = align_org

            # r0+1: Наименование - объединено на 2 колонки
            ws.merge_cells(start_row=r0 + 1, start_column=c0, end_row=r0 + 1, end_column=c0 + 1)
            cell_name = ws.cell(row=r0 + 1, column=c0)
            cell_name.value = wrapped_names[i]
            cell_name.font = name_font
            cell_name.alignment = align_name

            # r0+2: Цена - объединено на 2 колонки, динамический размер
            ws.merge_cells(start_row=r0 + 2, start_column=c0, end_row=r0 + 2, end_column=c0 + 1)
            cell_price = ws.cell(row=r0 + 2, column=c0)
            price_int = int(round(price))
            digits = len(str(price_int))
            price_font_size = price_base if digits <= 4 else (price_base - (digits - 4))
            if price_font_size < 12:
                price_font_size = 12
            cell_price.font = Font(name=font_name, size=price_font_size, bold=bool(price_bold))
            cell_price.value = price_int
            cell_price.alignment = align_price
            cell_price.number_format = '0 " р."'

            # r0+3: Дата - объединено на 2 колонки
            ws.merge_cells(start_row=r0 + 3, start_column=c0, end_row=r0 + 3, end_column=c0 + 1)
            cell_date = ws.cell(row=r0 + 3, column=c0)
            cell_date.value = f"Дата: {date_str}"
            cell_date.font = date_font
            cell_date.alignment = align_date

        # Параметры страницы
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = False

        # Сохраняем
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'ценники_{timestamp}.xlsx'
        filepath = os.path.join(CONVERTER_OUTPUT_DIR, filename)
        wb.save(filepath)

        # Метаданные
        meta = {
            'filename': filename,
            'created_at': datetime.now().isoformat(),
            'products_count': len(products),
            'settings': settings,
            'markup': markup
        }
        meta_path = os.path.join(CONVERTER_OUTPUT_DIR, f'{timestamp}.json')
        with open(meta_path, 'w', encoding='utf-8') as f:
            json.dump(meta, f, ensure_ascii=False, indent=2)

        logger.info(f"Price tags saved: {filepath}")

        return send_file(
            filepath,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        logger.exception(f"Generate error: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


@converter_bp.route('/converter/print', methods=['POST'])
def print_price_tags():
    """Печать ценников (возвращает HTML для печати)"""
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401
    
    data = request.json or {}
    tags = data.get('tags', [])
    
    # Генерируем HTML для печати
    html = '''
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <title>Ценники</title>
        <style>
            @page { size: A6; margin: 5mm; }
            body { font-family: Arial, sans-serif; }
            .tag { 
                border: 2px solid #000; 
                padding: 10px; 
                margin: 5px;
                page-break-inside: avoid;
            }
            .name { font-size: 14px; font-weight: bold; margin-bottom: 5px; }
            .price { font-size: 24px; font-weight: bold; color: #d00; }
            .unit { font-size: 12px; color: #666; }
        </style>
    </head>
    <body>
    '''
    
    for tag in tags:
        html += f'''
        <div class="tag">
            <div class="name">{tag['name']}</div>
            <div class="price">{tag['price']:.2f} ₽</div>
            <div class="unit">за {tag['unit']}</div>
        </div>
        '''
    
    html += '''
    </body>
    </html>
    '''
    
    from flask import make_response
    response = make_response(html)
    response.headers['Content-Type'] = 'text/html; charset=utf-8'
    return response


@converter_bp.route('/converter/generate-promotion', methods=['POST'])
def generate_promotion_price_tags():
    """
    Генерация акционных ценников в формате конвертера (толстые рамки, один столбик)
    с акционными элементами: старая/новая цена, "АКЦИЯ", огонёк
    """
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401

    data = request.json or {}
    products = data.get('products', [])
    settings = data.get('settings', {})
    org_name = data.get('org_name', 'ООО "ВетГид"')

    if not products:
        return jsonify({'status': 'error', 'message': 'Нет товаров'}), 400

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl import Workbook
        from math import ceil
        from datetime import datetime
    except ImportError:
        return jsonify({'status': 'error', 'message': 'openpyxl не установлен'}), 500

    # Создаём новую книгу
    wb = Workbook()
    ws = wb.active
    ws.title = "Акционные ценники"

    # Настройки из UI - используем формат конвертера
    n_cols = 1  # Один столбик как просил пользователь
    org_size = int(settings.get('org_size', 14))
    org_bold = settings.get('org_bold', True)
    name_size = int(settings.get('name_size', 14))
    name_bold = settings.get('name_bold', False)
    price_base = int(settings.get('price_base', 24))
    price_bold = settings.get('price_bold', True)
    date_size = int(settings.get('date_size', 10))
    date_bold = settings.get('date_bold', True)
    font_name = settings.get('font_name', 'Calibri')

    # Параметры как в конвертере
    TOTAL_COL_WIDTH = 35
    NAME_COL_WIDTH = 35
    PRICE_COL_WIDTH = 12
    LINE_HEIGHT_FACTOR = 1.35
    NAME_PADDING = 16
    MAX_NAME_LINES = 5

    # Текущая дата
    date_str = datetime.now().strftime('%d.%m.%Y')

    # Подготовка данных
    items = []
    for p in products:
        name = p.get('name', '')
        price = float(p.get('price', 0))
        old_price = float(p.get('old_price', 0)) if p.get('old_price') else 0
        unit = p.get('unit', 'шт')
        discount = p.get('discount', 0)
        items.append((name, price, old_price, unit, discount))

    total = len(items)

    # Блок ценника: 2 колонки x 5 строк (добавили строку для "АКЦИЯ")
    block_w = 2
    block_h = 5

    # Устанавливаем ширины колонок
    for col_block in range(n_cols):
        c_name = col_block * block_w + 1
        c_price = col_block * block_w + 2
        ws.column_dimensions[get_column_letter(c_name)].width = NAME_COL_WIDTH
        ws.column_dimensions[get_column_letter(c_price)].width = PRICE_COL_WIDTH

    # Границы как в конвертере
    thin = Side(style="thin", color="000000")
    medium = Side(style="medium", color="000000")

    # Выравнивания
    align_org = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_name = Alignment(horizontal="left", vertical="top", wrap_text=True)
    align_price = Alignment(horizontal="center", vertical="center", wrap_text=False, shrink_to_fit=True)
    align_date = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_promo = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Шрифты
    org_font = Font(name=font_name, size=org_size, bold=bool(org_bold))
    name_font = Font(name=font_name, size=name_size, bold=bool(name_bold))
    price_font = Font(name=font_name, size=price_base, bold=bool(price_bold), color='FFFF0000')  # Красный
    old_price_font = Font(name=font_name, size=price_base - 4, bold=False, color='FF808080')  # Серый
    date_font = Font(name=font_name, size=date_size, bold=bool(date_bold))
    promo_font = Font(name=font_name, size=14, bold=True, color='FFFFFFFF')  # Белый на красном

    # Фон для "АКЦИЯ"
    promo_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

    # Высоты строк
    org_row_height = int(org_size * LINE_HEIGHT_FACTOR + 6)
    promo_row_height = int(14 * LINE_HEIGHT_FACTOR + 8)
    date_row_height = int(date_size * LINE_HEIGHT_FACTOR + 6)

    # Функция переноса текста (из конвертера)
    def wrap_text_by_width(text: str, max_chars: int) -> str:
        text = str(text).strip()
        if not text:
            return ""
        if len(text) <= max_chars:
            return text

        words = text.split()
        if not words:
            return text

        lines = []
        current_line = []
        current_length = 0

        for word in words:
            word_len = len(word)
            if word_len > max_chars:
                if current_line:
                    lines.append(" ".join(current_line))
                    current_line = []
                    current_length = 0
                for i in range(0, len(word), max_chars):
                    chunk = word[i:i + max_chars]
                    if i + max_chars < len(word):
                        lines.append(chunk)
                    else:
                        current_line = [chunk]
                        current_length = len(chunk)
                continue

            if current_length == 0:
                current_line = [word]
                current_length = word_len
            elif current_length + 1 + word_len <= max_chars:
                current_line.append(word)
                current_length += 1 + word_len
            else:
                lines.append(" ".join(current_line))
                current_line = [word]
                current_length = word_len

        if current_line:
            lines.append(" ".join(current_line))

        return "\n".join(lines)

    # Перенос наименований
    name_max_chars = NAME_COL_WIDTH + PRICE_COL_WIDTH
    n_block_rows = ceil(total / n_cols) if total > 0 else 0
    max_lines_per_row = [1] * n_block_rows
    wrapped_names = [""] * total

    for i, (name, _price, _old_price, _unit, _discount) in enumerate(items):
        wrapped = wrap_text_by_width(name, name_max_chars)
        lines_count = max(1, wrapped.count("\n") + 1)
        if lines_count > MAX_NAME_LINES:
            lines = wrapped.split("\n")[:MAX_NAME_LINES]
            last_line = lines[-1]
            if len(last_line) > name_max_chars - 3:
                lines[-1] = last_line[:name_max_chars - 3] + "..."
            wrapped = "\n".join(lines)
            lines_count = MAX_NAME_LINES
        wrapped_names[i] = wrapped
        rb = i // n_cols
        if rb < len(max_lines_per_row) and lines_count > max_lines_per_row[rb]:
            max_lines_per_row[rb] = lines_count

    # Глобальный максимум строк
    global_max_lines = max(max_lines_per_row) if max_lines_per_row else 1
    for rb in range(n_block_rows):
        max_lines_per_row[rb] = global_max_lines

    # Устанавливаем высоты строк
    for rb in range(n_block_rows):
        r0 = rb * block_h + 1
        ws.row_dimensions[r0].height = org_row_height
        ws.row_dimensions[r0 + 1].height = promo_row_height  # Строка "АКЦИЯ"
        
        # Высота для наименования
        lines_count = max_lines_per_row[rb]
        text_height = name_size * LINE_HEIGHT_FACTOR * lines_count
        total_padding = NAME_PADDING * 2
        name_row_height = int(text_height + total_padding)
        min_height = int(name_size * LINE_HEIGHT_FACTOR + NAME_PADDING * 2)
        if name_row_height < min_height:
            name_row_height = min_height
        name_row_height = int(name_row_height * 1.1)
        ws.row_dimensions[r0 + 2].height = name_row_height
        
        # Цена и дата
        price_row_height = int(price_base * LINE_HEIGHT_FACTOR + 10)
        ws.row_dimensions[r0 + 3].height = price_row_height
        ws.row_dimensions[r0 + 4].height = date_row_height

    # Вывод блоков
    for i, (name, price, old_price, unit, discount) in enumerate(items):
        rb = i // n_cols
        cb = i % n_cols
        r0 = rb * block_h + 1
        c0 = cb * block_w + 1

        # Границы блока: внешние medium, внутренние thin
        for rr in range(r0, r0 + block_h):
            for cc in range(c0, c0 + block_w):
                left = medium if cc == c0 else thin
                right = medium if cc == c0 + block_w - 1 else thin
                top = medium if rr == r0 else thin
                bottom = medium if rr == r0 + block_h - 1 else thin
                ws.cell(row=rr, column=cc).border = Border(left=left, right=right, top=top, bottom=bottom)

        # r0: Организация (шапка) - объединено на 2 колонки
        ws.merge_cells(start_row=r0, start_column=c0, end_row=r0, end_column=c0 + 1)
        cell_org = ws.cell(row=r0, column=c0)
        cell_org.value = org_name
        cell_org.font = org_font
        cell_org.alignment = align_org

        # r0+1: "АКЦИЯ" с огоньком - объединено на 2 колонки
        ws.merge_cells(start_row=r0 + 1, start_column=c0, end_row=r0 + 1, end_column=c0 + 1)
        cell_promo = ws.cell(row=r0 + 1, column=c0)
        cell_promo.value = '🔥 АКЦИЯ'
        cell_promo.font = promo_font
        cell_promo.fill = promo_fill
        cell_promo.alignment = align_promo

        # r0+2: Наименование - объединено на 2 колонки
        ws.merge_cells(start_row=r0 + 2, start_column=c0, end_row=r0 + 2, end_column=c0 + 1)
        cell_name = ws.cell(row=r0 + 2, column=c0)
        cell_name.value = wrapped_names[i]
        cell_name.font = name_font
        cell_name.alignment = align_name

        # r0+3: Цена - объединено на 2 колонки
        ws.merge_cells(start_row=r0 + 3, start_column=c0, end_row=r0 + 3, end_column=c0 + 1)
        cell_price = ws.cell(row=r0 + 3, column=c0)
        
        # Форматируем цену: новая цена красная, старая серая с перечёркиванием
        if old_price and old_price > price:
            # Рассчитываем скидку
            discount_percent = int(((old_price - price) / old_price) * 100)
            # Форматируем цену: новая цена красная, старая серая
            price_text = f'{price:.2f} ₽\nбыло {old_price:.2f} ₽ (-{discount_percent}%)'
        else:
            price_text = f'{price:.2f} ₽'
        
        cell_price.value = price_text
        cell_price.font = price_font
        cell_price.alignment = align_price

        # r0+4: Дата и единица измерения - объединено на 2 колонки
        ws.merge_cells(start_row=r0 + 4, start_column=c0, end_row=r0 + 4, end_column=c0 + 1)
        cell_date = ws.cell(row=r0 + 4, column=c0)
        cell_date.value = f'{date_str} | за {unit}'
        cell_date.font = date_font
        cell_date.alignment = align_date

    # Параметры страницы
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = False

    # Сохраняем
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'акционные_ценники_{timestamp}.xlsx'
    filepath = os.path.join(CONVERTER_OUTPUT_DIR, filename)
    wb.save(filepath)

    # Метаданные
    meta = {
        'filename': filename,
        'created_at': datetime.now().isoformat(),
        'products_count': len(products),
        'settings': settings,
        'type': 'promotion'
    }
    meta_path = os.path.join(CONVERTER_OUTPUT_DIR, f'promotion_{timestamp}.json')
    with open(meta_path, 'w', encoding='utf-8') as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)

    logger.info(f"Promotion price tags generated: {len(products)} products, file: {filename}")

    # Отправляем файл
    return send_file(
        filepath,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
