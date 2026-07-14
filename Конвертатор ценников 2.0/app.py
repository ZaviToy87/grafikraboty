# app.py - Конвертер ценников 2.0 (обновленная версия)
import os
import sys
import subprocess
from datetime import datetime
from math import ceil
from typing import List, Optional, Tuple, Dict, Any

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox
    from tkinter import font as tkfont
except Exception as e:
    raise SystemExit(f"Ошибка инициализации Tkinter: {e}")

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Side, Font
    from openpyxl.worksheet.worksheet import Worksheet
except Exception as e:
    raise SystemExit(
        "Не установлены зависимости для работы с Excel. Установите пакет 'openpyxl':\n"
        "  pip install openpyxl\n\n"
        f"Подробности: {e}"
    )

ORG_NAME = 'ООО "КАКИЕ ЛЮДИ"'
DEFAULT_COLUMNS = 3
MAX_COLUMNS = 8

# Визуальные параметры (в условных ед. Excel)
TOTAL_COL_WIDTH = 35  # суммарная ширина ценника
NAME_COL_WIDTH = 26   # ширина колонки наименоования
PRICE_COL_WIDTH = 9   # ширина колонки цены

# Значения по умолчанию для шрифтов
DEFAULT_ORG_SIZE = 8
DEFAULT_ORG_BOLD = True
DEFAULT_NAME_SIZE = 10
DEFAULT_NAME_BOLD = False
DEFAULT_PRICE_BASE = 18
DEFAULT_PRICE_BOLD = True
DEFAULT_DATE_SIZE = 11
DEFAULT_DATE_BOLD = False

LINE_HEIGHT_FACTOR = 1.20
NAME_PADDING = 6

# Колонки 1С выгрузки (0-индексированные)
COLUMN_PRODUCT = "Товар"  # Наименование товара
COLUMN_PRICE = "Цена (руб)"  # Розничная цена


def detect_header_and_columns(ws: Worksheet, max_scan_rows: int = 20) -> Tuple[Optional[int], Optional[int], int]:
    """
    Пытается обнаружить строку заголовка и индексы столбцов для наименования и цены.
    Возвращает кортеж: (name_col_idx, price_col_idx, header_row_idx)
    Индексы 0-основанные. Если не найдены — возвращаются None, а header_row_idx=0.
    """
    # Ключевые слова для 1С выгрузки и других форматов
    name_keywords = ["товар", "наименование", "номенклатура", "product", "name", "наименование товара"]
    # Приоритет: новая розничная цена (колонка 9), затем старая розничная (колонка 7)
    price_keywords_new_retail = ["новая розничная цена", "новая розничная", "цена новая"]
    price_keywords_retail = ["розничная цена", "розничная"]

    found_header_row = None
    found_name_col = None
    found_new_retail_cols = []  # Может быть несколько колонок с "новая"
    found_retail_cols = []      # Может быть несколько колонок с "розничная"

    max_row_to_scan = min(ws.max_row, max_scan_rows)

    # Сканируем несколько строк в поисках заголовков
    for r in range(1, max_row_to_scan + 1):
        row_values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        row_texts = [(str(v).strip().lower() if v is not None else None) for v in row_values]

        name_candidates = [i for i, txt in enumerate(row_texts) if txt and any(k in txt for k in name_keywords)]
        
        # Ищем "новая розничная" / "цена новая" (приоритет - колонка 9)
        new_retail_candidates = [i for i, txt in enumerate(row_texts) if txt and any(k in txt for k in price_keywords_new_retail)]
        # Затем просто "розничная" (колонка 7)
        retail_candidates = [i for i, txt in enumerate(row_texts) if txt and any(k in txt for k in price_keywords_retail)]

        if name_candidates or new_retail_candidates or retail_candidates:
            if found_header_row is None:
                found_header_row = r
            if name_candidates and found_name_col is None:
                found_name_col = name_candidates[0]
            # Собираем все колонки с "новая розничная"
            found_new_retail_cols.extend(new_retail_candidates)
            # Собираем все колонки с "розничная"
            found_retail_cols.extend(retail_candidates)

    if found_header_row is None:
        found_header_row = 1

    # Если конкретные колонки не найдены в заголовке — ставим дефолты
    if found_name_col is None:
        found_name_col = 0  # Первый столбец для наименования
    
    # Приоритет: новая розничная цена (выбираем правую колонку - индекс 8 для колонки 9)
    # Если есть несколько колонок с "новая", выбираем самую правую (с большим индексом)
    if found_new_retail_cols:
        found_price_col = max(found_new_retail_cols)  # Правая колонка (колонка 9, индекс 8)
    elif found_retail_cols:
        found_price_col = max(found_retail_cols)  # Правая колонка с "розничная"
    else:
        found_price_col = 8  # 9-й столбец по умолчанию (новая розничная цена)

    return found_name_col, found_price_col, found_header_row


def read_items_from_excel(path: str) -> List[Tuple[str, float]]:
    """
    Читает Excel и возвращает список (наименование, цена).
    Использует эвристику для нахождения столбцов и начала данных.
    Пустые наименования/цены пропускаются.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    name_col, price_col, header_row = detect_header_and_columns(ws)

    items: List[Tuple[str, float]] = []

    for r in range(header_row + 1, ws.max_row + 1):
        name_val = ws.cell(row=r, column=name_col + 1).value
        price_val = ws.cell(row=r, column=price_col + 1).value

        if name_val is None:
            continue

        name_str = str(name_val).strip()
        if not name_str:
            continue

        # Преобразуем цену в float если возможно
        price_num: Optional[float] = None
        if price_val is not None and str(price_val).strip() != "":
            try:
                price_num = float(str(price_val).replace(" ", "").replace(",", "."))
            except Exception:
                price_num = None

        if price_num is None:
            continue

        items.append((name_str, price_num))

    return items


def format_price(price: float) -> int:
    # Округляем до целых и возвращаем число (без символов и разделителей)
    return int(round(price))


def wrap_text_by_width(text: str, max_chars: int) -> str:
    """Перенос строк только по словам так, чтобы длина строки не превышала max_chars."""
    words = str(text).split()
    if max_chars <= 0:
        return " ".join(words)
    lines = []
    line = ""
    for w in words:
        if not line:
            line = w
        elif len(line) + 1 + len(w) <= max_chars:
            line += " " + w
        else:
            lines.append(line)
            line = w
    if line:
        lines.append(line)
    return "\n".join(lines)


def generate_tags_excel(
    items: List[Tuple[str, float]],
    n_cols: int,
    out_path: str,
    fonts_conf: Optional[Dict[str, Any]] = None,
) -> None:
    """Формирует Excel с ценниками по заданным настройкам шрифтов."""
    # Настройки по умолчанию
    fc = {
        "org_name": ORG_NAME,
        "org_size": DEFAULT_ORG_SIZE,
        "org_bold": DEFAULT_ORG_BOLD,
        "name_size": DEFAULT_NAME_SIZE,
        "name_bold": DEFAULT_NAME_BOLD,
        "price_base": DEFAULT_PRICE_BASE,
        "price_bold": DEFAULT_PRICE_BOLD,
        "date_size": DEFAULT_DATE_SIZE,
        "date_bold": DEFAULT_DATE_BOLD,
    }
    if fonts_conf:
        fc.update(fonts_conf)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ценники"

    date_str = datetime.now().strftime("%d.%m.%Y")

    # Макет одного ценника: 2 колонки x 4 строки (шапка, наименование, цена (всю ширину), дата)
    block_w = 2
    block_h = 4

    total = len(items)

    # Устанавливаем ширины колонок для каждого блока в ряду
    for col_block in range(n_cols):
        c_name = col_block * block_w + 1
        c_price = col_block * block_w + 2
        ws.column_dimensions[ws.cell(row=1, column=c_name).column_letter].width = NAME_COL_WIDTH
        ws.column_dimensions[ws.cell(row=1, column=c_price).column_letter].width = PRICE_COL_WIDTH

    thin = Side(style="thin", color="000000")
    medium = Side(style="medium", color="000000")

    align_org = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_name = Alignment(horizontal="left", vertical="top", wrap_text=True)
    align_price = Alignment(horizontal="center", vertical="center", wrap_text=False, shrink_to_fit=True)
    align_date = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Подготовка переносов наименования и расчёт высоты строки (r0+1) по ряду
    # Теперь имя занимает всю ширину блока (две колонки), поэтому ширина для расчёта переносов увеличена
    name_max_chars = NAME_COL_WIDTH + PRICE_COL_WIDTH  # эвристика: используем суммарную ширину блока
    n_block_rows = ceil(total / n_cols) if total > 0 else 0
    max_lines_per_row = [1] * n_block_rows
    wrapped_names: List[str] = [""] * total

    for i, (name, _price) in enumerate(items):
        wrapped = wrap_text_by_width(name, name_max_chars)
        wrapped_names[i] = wrapped
        lines_count = max(1, wrapped.count("\n") + 1)
        rb = i // n_cols
        if rb < len(max_lines_per_row) and lines_count > max_lines_per_row[rb]:
            max_lines_per_row[rb] = lines_count

    # Локальные шрифты/высоты строк из настроек
    org_font = Font(name="Calibri", size=fc["org_size"], bold=bool(fc["org_bold"]))
    name_font = Font(name="Calibri", size=fc["name_size"], bold=bool(fc["name_bold"]))
    date_font = Font(name="Calibri", size=fc["date_size"], bold=bool(fc["date_bold"]))

    org_row_height = int(fc["org_size"] * LINE_HEIGHT_FACTOR + 6)
    date_row_height = int(fc["date_size"] * LINE_HEIGHT_FACTOR + 6)

    # Высоты строк: шапка фикс., строка с наименованием завязана на макс. число строк,
    # строка с ценой — завязана на размер шрифта цены, дата фикс.
    for rb in range(n_block_rows):
        r0 = rb * block_h + 1
        ws.row_dimensions[r0].height = org_row_height
        name_row_height = int(fc["name_size"] * LINE_HEIGHT_FACTOR * max_lines_per_row[rb] + NAME_PADDING)
        ws.row_dimensions[r0 + 1].height = name_row_height
        # Цена: базовый размер шрифта -> высота строки
        price_row_height = int(fc["price_base"] * LINE_HEIGHT_FACTOR + 6)
        ws.row_dimensions[r0 + 2].height = price_row_height
        ws.row_dimensions[r0 + 3].height = date_row_height

    # Выводим блоки
    for i, (name, price) in enumerate(items):
        rb = i // n_cols
        cb = i % n_cols
        r0 = rb * block_h + 1
        c0 = cb * block_w + 1

        # Границы блока: внешняя средняя, внутренние тонкие
        for rr in range(r0, r0 + block_h):
            for cc in range(c0, c0 + block_w):
                left = medium if cc == c0 else thin
                right = medium if cc == c0 + block_w - 1 else thin
                top = medium if rr == r0 else thin
                bottom = medium if rr == r0 + block_h - 1 else thin
                ws.cell(row=rr, column=cc).border = Border(left=left, right=right, top=top, bottom=bottom)

        # r0: шапка (организация) — объединено
        ws.merge_cells(start_row=r0, start_column=c0, end_row=r0, end_column=c0 + 1)
        cell_org = ws.cell(row=r0, column=c0)
        cell_org.value = fc["org_name"]
        cell_org.font = org_font
        cell_org.alignment = align_org

        # r0+1: наименование — теперь объединено по всей ширине блока
        ws.merge_cells(start_row=r0 + 1, start_column=c0, end_row=r0 + 1, end_column=c0 + 1)
        cell_name = ws.cell(row=r0 + 1, column=c0)
        cell_name.value = wrapped_names[i]
        cell_name.font = name_font
        cell_name.alignment = align_name

        # r0+2: цена — объединено по всей ширине блока, крупный шрифт, по центру
        ws.merge_cells(start_row=r0 + 2, start_column=c0, end_row=r0 + 2, end_column=c0 + 1)
        cell_price = ws.cell(row=r0 + 2, column=c0)
        price_int = format_price(price)
        digits = len(str(price_int))
        price_font_size = fc["price_base"] if digits <= 4 else (fc["price_base"] - (digits - 4))
        if price_font_size < 12:
            price_font_size = 12
        cell_price.font = Font(name="Calibri", size=price_font_size, bold=bool(fc["price_bold"]))
        cell_price.value = price_int
        cell_price.alignment = align_price
        cell_price.number_format = '0 " р."'

        # r0+3: дата (сегодня), объединено на 2 колонки
        ws.merge_cells(start_row=r0 + 3, start_column=c0, end_row=r0 + 3, end_column=c0 + 1)
        cell_date = ws.cell(row=r0 + 3, column=c0)
        cell_date.value = f"Дата: {date_str}"
        cell_date.font = date_font
        cell_date.alignment = align_date

    # Параметры страницы для печати
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = False

    wb.save(out_path)


def get_desktop_files_folder() -> str:
    """Получить путь к папке 'Файлы' на рабочем столе. Создать если нет."""
    try:
        import ctypes.windll
        from ctypes import wintypes

        CSIDL_DESKTOPDIRECTORY = 0x0010
        SHGFP_TYPE_CURRENT = 0

        buf = ctypes.create_unicode_buffer(ctypes.wintypes.MAX_PATH)
        ctypes.windll.shell32.SHGetFolderPathW(None, CSIDL_DESKTOPDIRECTORY, None, SHGFP_TYPE_CURRENT, buf)
        desktop_path = buf.value
    except Exception:
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")

    files_folder = os.path.join(desktop_path, "Файлы")

    if not os.path.exists(files_folder):
        try:
            os.makedirs(files_folder)
        except Exception:
            files_folder = desktop_path

    return files_folder


def get_created_files_list() -> List[str]:
    """Получить список созданных файлов ценников в папке 'Файлы'."""
    files_folder = get_desktop_files_folder()
    if not os.path.exists(files_folder):
        return []
    
    files = []
    for f in os.listdir(files_folder):
        if f.startswith("ценники_") and f.endswith(".xlsx"):
            files.append(os.path.join(files_folder, f))
    
    # Сортируем по дате изменения (новые сверху)
    files.sort(key=lambda x: os.path.getmtime(x), reverse=True)
    return files


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Конвертер ценников 2.0")
        self.geometry("1200x650")
        self.resizable(True, True)

        # Путь к входному файлу
        self.input_path_var = tk.StringVar(value=self._default_input_path())

        # Количество столбцов
        self.cols_var = tk.IntVar(value=DEFAULT_COLUMNS)

        # Настройки шрифтов/отображения
        self.org_name_var = tk.StringVar(value=ORG_NAME)
        self.org_size_var = tk.IntVar(value=DEFAULT_ORG_SIZE)
        self.org_bold_var = tk.BooleanVar(value=DEFAULT_ORG_BOLD)

        self.name_size_var = tk.IntVar(value=DEFAULT_NAME_SIZE)
        self.name_bold_var = tk.BooleanVar(value=DEFAULT_NAME_BOLD)

        self.price_base_var = tk.IntVar(value=DEFAULT_PRICE_BASE)
        self.price_bold_var = tk.BooleanVar(value=DEFAULT_PRICE_BOLD)

        self.date_size_var = tk.IntVar(value=DEFAULT_DATE_SIZE)
        self.date_bold_var = tk.BooleanVar(value=DEFAULT_DATE_BOLD)

        # Данные из Excel для превью/выгрузки
        self.items: List[Tuple[str, float]] = []

        # Для списка файлов
        self.files_listbox = None
        self.files_scrollbar = None

        self._build_ui()
        self.after(50, self.update_preview)
        self.after(100, self.refresh_files_list)

    def _default_input_path(self) -> str:
        # Предполагаемый файл по умолчанию в рабочем каталоге
        cwd = os.getcwd()
        potential = os.path.join(cwd, "1111.xlsx")
        return potential if os.path.exists(potential) else ""

    def _build_ui(self):
        padx = 12
        pady = 8

        root = tk.Frame(self)
        root.pack(fill=tk.BOTH, expand=True, padx=padx, pady=pady)

        # Три колонки: слева панель файлов, центр настройки, справа превью
        left_panel = tk.Frame(root, width=250)
        left_panel.grid(row=0, column=0, sticky="nsew")
        left_panel.grid_propagate(False)

        center = tk.Frame(root)
        center.grid(row=0, column=1, sticky="nsew", padx=(12, 12))

        right = tk.Frame(root)
        right.grid(row=0, column=2, sticky="nsew")

        root.grid_columnconfigure(0, weight=0)
        root.grid_columnconfigure(1, weight=0)
        root.grid_columnconfigure(2, weight=1)
        root.grid_rowconfigure(0, weight=1)

        # --------- LEFT PANEL: Список созданных файлов ---------
        files_frame = tk.LabelFrame(left_panel, text="Созданные файлы")
        files_frame.pack(fill=tk.BOTH, expand=True)

        # Список файлов
        list_frame = tk.Frame(files_frame)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)

        self.files_scrollbar = tk.Scrollbar(list_frame)
        self.files_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.files_listbox = tk.Listbox(
            list_frame,
            width=30,
            height=20,
            yscrollcommand=self.files_scrollbar.set,
            font=("Segoe UI", 9)
        )
        self.files_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.files_scrollbar.config(command=self.files_listbox.yview)

        # Привязка двойного клика
        self.files_listbox.bind("<Double-Button-1>", self.on_file_double_click)

        # Кнопки действий с файлами
        btn_frame = tk.Frame(files_frame)
        btn_frame.pack(fill=tk.X, padx=4, pady=(0, 4))

        btn_open = tk.Button(btn_frame, text="Открыть", command=self.on_open_file, width=12)
        btn_open.pack(side=tk.LEFT, padx=2, pady=4)

        btn_print = tk.Button(btn_frame, text="Печать", command=self.on_print_file, width=12)
        btn_print.pack(side=tk.LEFT, padx=2, pady=4)

        btn_refresh = tk.Button(btn_frame, text="Обновить", command=self.refresh_files_list, width=12)
        btn_refresh.pack(side=tk.LEFT, padx=2, pady=4)

        # --------- CENTER: файл и общие параметры ---------
        tk.Label(center, text="Файл Excel с товарами:").grid(row=0, column=0, columnspan=3, sticky="w")

        entry = tk.Entry(center, textvariable=self.input_path_var, width=64)
        entry.grid(row=1, column=0, columnspan=2, sticky="we")

        btn_browse = tk.Button(center, text="Загрузить файл", command=self.on_browse)
        btn_browse.grid(row=1, column=2, sticky="we", padx=(8, 0))

        tk.Label(center, text="Количество столбцов ценников:").grid(row=2, column=0, sticky="w", pady=(16, 0))
        spin_cols = tk.Spinbox(center, from_=1, to=MAX_COLUMNS, textvariable=self.cols_var, width=5, command=self.update_preview)
        spin_cols.grid(row=2, column=1, sticky="w", pady=(16, 0))

        # Настройки шрифтов
        lf = tk.LabelFrame(center, text="Настройки отображения")
        lf.grid(row=3, column=0, columnspan=3, sticky="we", pady=(16, 0))

        # Организация
        tk.Label(lf, text="Организация (шапка):").grid(row=0, column=0, sticky="w", pady=(6, 0))
        org_entry = tk.Entry(lf, textvariable=self.org_name_var, width=40)
        org_entry.grid(row=0, column=1, columnspan=2, sticky="we", pady=(6, 0))
        tk.Label(lf, text="Размер").grid(row=1, column=0, sticky="w")
        tk.Spinbox(lf, from_=6, to=48, textvariable=self.org_size_var, width=5, command=self.update_preview).grid(row=1, column=1, sticky="w")
        tk.Checkbutton(lf, text="Жирный", variable=self.org_bold_var, command=self.update_preview).grid(row=1, column=2, sticky="w")

        # Наименование
        tk.Label(lf, text="Наименование:").grid(row=2, column=0, sticky="w", pady=(10, 0))
        tk.Label(lf, text="Размер").grid(row=3, column=0, sticky="w")
        tk.Spinbox(lf, from_=6, to=48, textvariable=self.name_size_var, width=5, command=self.update_preview).grid(row=3, column=1, sticky="w")
        tk.Checkbutton(lf, text="Жирный", variable=self.name_bold_var, command=self.update_preview).grid(row=3, column=2, sticky="w")

        # Цена
        tk.Label(lf, text="Цена (базовый размер):").grid(row=4, column=0, sticky="w", pady=(10, 0))
        tk.Label(lf, text="Размер").grid(row=5, column=0, sticky="w")
        tk.Spinbox(lf, from_=10, to=64, textvariable=self.price_base_var, width=5, command=self.update_preview).grid(row=5, column=1, sticky="w")
        tk.Checkbutton(lf, text="Жирный", variable=self.price_bold_var, command=self.update_preview).grid(row=5, column=2, sticky="w")

        # Дата
        tk.Label(lf, text="Дата:").grid(row=6, column=0, sticky="w", pady=(10, 0))
        tk.Label(lf, text="Размер").grid(row=7, column=0, sticky="w")
        tk.Spinbox(lf, from_=6, to=48, textvariable=self.date_size_var, width=5, command=self.update_preview).grid(row=7, column=1, sticky="w")
        tk.Checkbutton(lf, text="Жирный", variable=self.date_bold_var, command=self.update_preview).grid(row=7, column=2, sticky="w")

        # Кнопка запуска
        btn_run = tk.Button(center, text="Сформировать ценники", command=self.on_run, height=2)
        btn_run.grid(row=8, column=0, columnspan=3, sticky="we", pady=(20, 0))

        # Логи/подсказки
        self.log_var = tk.StringVar(value="Подсказка: загрузите Excel, настройте внешний вид и нажмите Сформировать.")
        tk.Label(center, textvariable=self.log_var, fg="#333").grid(row=9, column=0, columnspan=3, sticky="w", pady=(16, 0))

        for i in range(3):
            center.grid_columnconfigure(i, weight=1)

        # --------- RIGHT: предпросмотр ---------
        preview_frame = tk.LabelFrame(right, text="Предпросмотр ценника (1 шт)")
        preview_frame.pack(fill=tk.BOTH, expand=True)

        self.preview_canvas = tk.Canvas(preview_frame, width=520, height=280, bg="#fff")
        self.preview_canvas.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

    def refresh_files_list(self):
        """Обновить список созданных файлов."""
        if self.files_listbox is None:
            return
        
        self.files_listbox.delete(0, tk.END)
        files = get_created_files_list()
        
        for f in files:
            filename = os.path.basename(f)
            # Добавляем дату создания
            try:
                mtime = os.path.getmtime(f)
                date_str = datetime.fromtimestamp(mtime).strftime("%d.%m.%Y %H:%M")
                self.files_listbox.insert(tk.END, f"{filename}\n  {date_str}")
            except Exception:
                self.files_listbox.insert(tk.END, filename)

    def on_file_double_click(self, event):
        """Обработка двойного клика по файлу."""
        self.on_open_file()

    def on_open_file(self):
        """Открыть выбранный файл в Excel."""
        if self.files_listbox is None:
            return
        
        selection = self.files_listbox.curselection()
        if not selection:
            messagebox.showwarning("Нет файла", "Выберите файл из списка.")
            return
        
        selected_text = self.files_listbox.get(selection[0])
        filename = selected_text.split('\n')[0]  # Берём первую строку (имя файла)
        
        files_folder = get_desktop_files_folder()
        file_path = os.path.join(files_folder, filename)
        
        if not os.path.exists(file_path):
            messagebox.showerror("Ошибка", f"Файл не найден:\n{file_path}")
            return
        
        try:
            # Открываем в Excel
            os.startfile(file_path)
        except Exception as e:
            messagebox.showerror("Ошибка открытия", f"Не удалось открыть файл:\n{e}")

    def on_print_file(self):
        """Отправить выбранный файл на печать."""
        if self.files_listbox is None:
            return
        
        selection = self.files_listbox.curselection()
        if not selection:
            messagebox.showwarning("Нет файла", "Выберите файл из списка.")
            return
        
        selected_text = self.files_listbox.get(selection[0])
        filename = selected_text.split('\n')[0]  # Берём первую строку (имя файла)
        
        files_folder = get_desktop_files_folder()
        file_path = os.path.join(files_folder, filename)
        
        if not os.path.exists(file_path):
            messagebox.showerror("Ошибка", f"Файл не найден:\n{file_path}")
            return
        
        try:
            # Печать через Excel
            excel_app = subprocess.Popen(
                ["excel", "/x", file_path],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE
            )
            # Ждём немного пока Excel откроется
            import time
            time.sleep(2)
            
            # Отправляем команду печати
            subprocess.Popen(
                ["powershell", "-Command", f"""
                    $excel = New-Object -ComObject Excel.Application
                    $workbook = $excel.Workbooks.Open('{file_path}')
                    $workbook.PrintOut()
                    $workbook.Close($false)
                    $excel.Quit()
                """],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE
            )
            
            self.log_var.set(f"Файл отправлен на печать: {filename}")
        except Exception as e:
            messagebox.showerror("Ошибка печати", f"Не удалось отправить файл на печать:\n{e}")

    def on_browse(self):
        path = filedialog.askopenfilename(
            title="Выберите Excel файл",
            filetypes=[("Excel файлы", "*.xlsx"), ("Все файлы", "*.*")],
            initialdir=os.getcwd(),
        )
        if not path:
            return
        self.input_path_var.set(path)
        # После выбора сразу пытаемся прочитать и показать превью
        try:
            self.log_var.set("Чтение данных из Excel...")
            self.update_idletasks()
            items = read_items_from_excel(path)
            if items:
                self.items = items
                self.log_var.set(f"Загружено позиций: {len(items)}. Превью обновлено.")
            else:
                self.items = []
                self.log_var.set("В файле не найдено валидных позиций. Используется демонстрационный пример для превью.")
        except Exception as e:
            self.items = []
            self.log_var.set(f"Ошибка чтения Excel: {e}")
            messagebox.showerror("Ошибка чтения Excel", f"Не удалось прочитать файл:\n{e}")
        finally:
            self.update_preview()

    def _collect_font_settings(self) -> Dict[str, Any]:
        return {
            "org_name": self.org_name_var.get().strip() or ORG_NAME,
            "org_size": int(self.org_size_var.get()),
            "org_bold": bool(self.org_bold_var.get()),
            "name_size": int(self.name_size_var.get()),
            "name_bold": bool(self.name_bold_var.get()),
            "price_base": int(self.price_base_var.get()),
            "price_bold": bool(self.price_bold_var.get()),
            "date_size": int(self.date_size_var.get()),
            "date_bold": bool(self.date_bold_var.get()),
        }

    def update_preview(self):
        # Подготовка данных для превью: берём первую позицию или демо
        if self.items:
            name, price = self.items[0]
        else:
            name = "Пример товара с длинным наименованием для проверки переноса строк"
            price = 12345

        fc = self._collect_font_settings()

        # Отрисовка одного блока ценника
        c = self.preview_canvas
        c.delete("all")

        # Размеры превью (пиксели)
        pad = 10
        W = int(c.winfo_width() or 520) - pad * 2
        H = int(c.winfo_height() or 280) - pad * 2
        x0, y0 = pad, pad

        # Теперь имя и цена занимают всю ширину блока
        total_w = W

        # Высота строк по размерам шрифтов и содержимому
        f_org = tkfont.Font(family="Calibri", size=fc["org_size"], weight=("bold" if fc["org_bold"] else "normal"))
        f_name = tkfont.Font(family="Calibri", size=fc["name_size"], weight=("bold" if fc["name_bold"] else "normal"))
        # Цена: динамическая регулировка от количества цифр (увеличиваем визуально)
        price_int = format_price(float(price))
        digits = len(str(price_int))
        dyn_price_size = fc["price_base"] if digits <= 4 else (fc["price_base"] - (digits - 4))
        if dyn_price_size < 12:
            dyn_price_size = 12
        # Немного увеличим базовый размер визуально в превью для акцента
        dyn_price_size = int(dyn_price_size * 1.15)
        f_price = tkfont.Font(family="Calibri", size=dyn_price_size, weight=("bold" if fc["price_bold"] else "normal"))
        f_date = tkfont.Font(family="Calibri", size=fc["date_size"], weight=("bold" if fc["date_bold"] else "normal"))

        # Перенос наименования по ширине пикселей — используем всю ширину
        name_lines = self._wrap_by_pixels(name, total_w - 8, f_name)

        org_h = int(fc["org_size"] * LINE_HEIGHT_FACTOR + 6)
        name_h = int(fc["name_size"] * LINE_HEIGHT_FACTOR * max(1, len(name_lines)) + NAME_PADDING)
        price_h = int(dyn_price_size * LINE_HEIGHT_FACTOR + 8)
        date_h = int(fc["date_size"] * LINE_HEIGHT_FACTOR + 6)

        total_h = org_h + name_h + price_h + date_h
        # Центрируем блок по вертикали в доступном H
        y0 = pad + max(0, (H - total_h) // 2)

        # Рамки
        # Внешняя рамка (без вертикальной разделительной линии — цена занимает всю ширину)
        c.create_rectangle(x0, y0, x0 + total_w, y0 + total_h, outline="#000", width=2)
        # Горизонтальные линии
        c.create_line(x0, y0 + org_h, x0 + total_w, y0 + org_h, fill="#000", width=1)
        c.create_line(x0, y0 + org_h + name_h, x0 + total_w, y0 + org_h + name_h, fill="#000", width=1)
        c.create_line(x0, y0 + org_h + name_h + price_h, x0 + total_w, y0 + org_h + name_h + price_h, fill="#000", width=1)

        # Тексты
        # Шапка — по центру
        c.create_text(x0 + total_w // 2, y0 + org_h // 2, text=fc["org_name"], font=f_org, anchor="c")

        # Наименование — слева, сверху (занимает всю ширину)
        ty = y0 + org_h + 4
        for i, line in enumerate(name_lines):
            c.create_text(x0 + 4, ty + i * int(fc["name_size"] * LINE_HEIGHT_FACTOR), text=line, font=f_name, anchor="nw")

        # Цена — по центру на своей строке, крупно
        cx = x0 + total_w // 2
        cy = y0 + org_h + name_h + price_h // 2
        c.create_text(cx, cy, text=f"{price_int}  р.", font=f_price, anchor="c")

        # Дата — слева внизу
        date_str = datetime.now().strftime("%d.%m.%Y")
        c.create_text(x0 + 4, y0 + org_h + name_h + price_h + date_h // 2, text=f"Дата: {date_str}", font=f_date, anchor="w")

    def _wrap_by_pixels(self, text: str, max_w: int, font: tkfont.Font) -> List[str]:
        words = str(text).split()
        if not words:
            return [""]
        lines: List[str] = []
        line = ""
        for w in words:
            if not line:
                candidate = w
            else:
                candidate = line + " " + w
            if font.measure(candidate) <= max_w:
                line = candidate
            else:
                if line:
                    lines.append(line)
                # слово само может быть длиннее ширины — жёстко переносим
                if font.measure(w) <= max_w:
                    line = w
                else:
                    # Разбивка по символам, пока не влезет
                    chunk = ""
                    for ch in w:
                        if font.measure(chunk + ch) <= max_w:
                            chunk += ch
                        else:
                            if chunk:
                                lines.append(chunk)
                            chunk = ch
                    line = chunk
        if line:
            lines.append(line)
        return lines

    def on_run(self):
        path = self.input_path_var.get().strip()
        if not path:
            messagebox.showwarning("Нет файла", "Укажите путь к Excel-файлу.")
            return
        if not os.path.exists(path):
            messagebox.showerror("Файл не найден", f"Файл не найден:\n{path}")
            return

        try:
            n_cols = int(self.cols_var.get())
        except Exception:
            messagebox.showerror("Ошибка", "Некорректное число столбцов.")
            return

        if n_cols < 1 or n_cols > MAX_COLUMNS:
            messagebox.showerror("Ошибка", f"Число столбцов должно быть от 1 до {MAX_COLUMNS}.")
            return

        # Если ранее не загружали данные — читаем сейчас
        if not self.items:
            self.log_var.set("Чтение данных из Excel...")
            self.update_idletasks()
            try:
                self.items = read_items_from_excel(path)
            except Exception as e:
                messagebox.showerror("Ошибка чтения Excel", f"Не удалось прочитать файл:\n{e}")
                return

        if not self.items:
            messagebox.showwarning("Нет данных", "Не найдено валидных товаров с ценой.")
            return

        # Сохраняем в папку 'Файлы' на рабочем столе
        out_dir = get_desktop_files_folder()
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_name = f"ценники_{ts}.xlsx"
        out_path = os.path.join(out_dir, out_name)

        self.log_var.set("Формирование Excel с ценниками...")
        self.update_idletasks()

        try:
            generate_tags_excel(self.items, n_cols, out_path, fonts_conf=self._collect_font_settings())
        except Exception as e:
            messagebox.showerror("Ошибка генерации", f"Не удалось сформировать ценники:\n{e}")
            return

        # Обновляем список файлов
        self.refresh_files_list()

        self.log_var.set(f"Готово: {out_path}")
        messagebox.showinfo("Готово", f"Файл создан:\n{out_path}\n\nФайл сохранён в папке:\n{out_dir}")


def main():
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
