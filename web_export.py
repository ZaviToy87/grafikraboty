# -*- coding: utf-8 -*-
"""
web_export.py — Экспорт отчётов в PDF/Excel
API: /api/export/*
"""
from flask import Blueprint, request, jsonify, session, send_file
from web_config import get_db_connection, logger
from datetime import datetime, timedelta
import os
import json
import tempfile

export_bp = Blueprint('export', __name__, url_prefix='/api/export')


def generate_excel_report(data, filename):
    """Генерация Excel отчёта"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Отчёт"

        # Заголовки
        headers = data.get('headers', [])
        rows = data.get('rows', [])

        # Стили
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Заголовок отчёта
        if data.get('title'):
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers) or 1)
            title_cell = ws.cell(row=1, column=1, value=data['title'])
            title_cell.font = Font(bold=True, size=14, color='6366F1')
            title_cell.alignment = Alignment(horizontal='center')
            start_row = 3
        else:
            start_row = 1

        # Записываем заголовки
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Записываем данные
        for row_idx, row_data in enumerate(rows, start_row + 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')

        # Авто-ширина колонок
        for col in range(1, len(headers) + 1):
            max_length = len(str(headers[col - 1]))
            for row in ws.iter_rows(min_col=col, max_col=col, values_only=True):
                for cell_val in row:
                    if cell_val:
                        max_length = max(max_length, len(str(cell_val)))
            ws.column_dimensions[chr(64 + col)].width = min(max_length + 2, 50)

        # Сохраняем во временный файл
        temp_dir = tempfile.gettempdir()
        filepath = os.path.join(temp_dir, filename)
        wb.save(filepath)
        return filepath

    except ImportError:
        logger.error("openpyxl not installed. Install with: pip install openpyxl")
        return None
    except Exception as e:
        logger.error(f"Error generating Excel: {e}")
        return None


def generate_pdf_report(data, filename):
    """Генерация PDF отчёта"""
    try:
        from weasyprint import HTML
        import tempfile

        # Строим HTML таблицу
        headers = data.get('headers', [])
        rows = data.get('rows', [])

        html_parts = ['<!DOCTYPE html><html><head><meta charset="utf-8">']
        html_parts.append('<style>')
        html_parts.append('''
            body { font-family: 'DejaVu Sans', Arial, sans-serif; font-size: 12px; }
            h1 { color: #6366f1; text-align: center; }
            table { width: 100%; border-collapse: collapse; margin-top: 20px; }
            th { background: #6366f1; color: white; padding: 8px; text-align: center; }
            td { padding: 6px; border: 1px solid #ddd; text-align: center; }
            tr:nth-child(even) { background: #f9f9f9; }
            .footer { text-align: center; margin-top: 20px; color: #888; font-size: 10px; }
        ''')
        html_parts.append('</style></head><body>')

        if data.get('title'):
            html_parts.append(f'<h1>{data["title"]}</h1>')

        html_parts.append('<table><thead><tr>')
        for h in headers:
            html_parts.append(f'<th>{h}</th>')
        html_parts.append('</tr></thead><tbody>')

        for row in rows:
            html_parts.append('<tr>')
            for cell in row:
                html_parts.append(f'<td>{cell}</td>')
            html_parts.append('</tr>')

        html_parts.append('</tbody></table>')
        html_parts.append(f'<div class="footer">Сгенерировано: {datetime.now().strftime("%d.%m.%Y %H:%M")}</div>')
        html_parts.append('</body></html>')

        html_content = '\n'.join(html_parts)

        # Генерируем PDF
        temp_dir = tempfile.gettempdir()
        filepath = os.path.join(temp_dir, filename)
        HTML(string=html_content).write_pdf(filepath)
        return filepath

    except ImportError:
        logger.error("weasyprint not installed. Install with: pip install weasyprint")
        return None
    except Exception as e:
        logger.error(f"Error generating PDF: {e}")
        return None


@export_bp.route('/schedule', methods=['POST'])
def export_schedule():
    """Экспорт графика работы"""
    if 'user_id' not in session:
        return jsonify({'error': 'Не авторизован'}), 401

    data = request.json or {}
    fmt = data.get('format', 'excel')
    year = data.get('year', datetime.now().year)
    month = data.get('month', datetime.now().month)

    try:
        db = get_db_connection()
        cursor = db.cursor()

        # Получаем все записи графика за месяц
        cursor.execute('''
            SELECT s.*, u.full_name
            FROM schedule s
            JOIN users u ON s.user_id = u.id
            WHERE s.year = ? AND s.month = ?
              AND s.task_ids != '[]' AND s.task_ids IS NOT NULL
            ORDER BY u.full_name, s.day
        ''', (year, month))

        rows = cursor.fetchall()

        # Получаем все задачи для маппинга id -> name
        cursor.execute('SELECT id, name, color FROM tasks')
        tasks_map = {row['id']: row for row in cursor.fetchall()}
        db.close()

        # Группируем по сотрудникам
        from collections import defaultdict
        schedule_data = defaultdict(lambda: defaultdict(list))

        for row in rows:
            name = row['full_name'] or f"User"
            day = row['day']
            # Парсим task_ids из JSON строки
            try:
                task_ids = json.loads(row['task_ids']) if row['task_ids'] else []
            except (json.JSONDecodeError, TypeError):
                task_ids = []
            if not isinstance(task_ids, list):
                task_ids = []
            task_names = []
            for tid in task_ids:
                task_info = tasks_map.get(int(tid)) if isinstance(tid, (int, str)) and str(tid).isdigit() else None
                if task_info:
                    task_names.append(task_info['name'])
            if not task_names:
                task_names = ['Смена']
            schedule_data[name][day] = task_names

        import calendar
        days_in_month = calendar.monthrange(year, month)[1]

        headers = ['Сотрудник'] + [str(d) for d in range(1, days_in_month + 1)]
        export_rows = []

        for emp_name, days in schedule_data.items():
            row_data = [emp_name]
            for d in range(1, days_in_month + 1):
                tasks = days.get(d, [])
                row_data.append(', '.join(tasks) if tasks else '')
            export_rows.append(row_data)

        month_names = ['', 'Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
                       'Июль', 'Август', 'Сентябрь', 'Окторябрь', 'Ноябрь', 'Декабрь']

        report_data = {
            'title': f'График работы — {month_names[month]} {year}',
            'headers': headers,
            'rows': export_rows
        }

        if fmt == 'excel':
            filename = f'schedule_{year}_{month:02d}.xlsx'
            filepath = generate_excel_report(report_data, filename)
            if filepath:
                return send_file(filepath, as_attachment=True, download_name=filename,
                               mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            return jsonify({'error': 'Ошибка генерации Excel. Установите openpyxl: pip install openpyxl'}), 500

        elif fmt == 'pdf':
            filename = f'schedule_{year}_{month:02d}.pdf'
            filepath = generate_pdf_report(report_data, filename)
            if filepath:
                return send_file(filepath, as_attachment=True, download_name=filename,
                               mimetype='application/pdf')
            return jsonify({'error': 'Ошибка генерации PDF. Установите weasyprint: pip install weasyprint'}), 500

        return jsonify({'error': 'Неподдерживаемый формат'}), 400

    except Exception as e:
        logger.error(f"Error exporting schedule: {e}")
        return jsonify({'error': str(e)}), 500


@export_bp.route('/revision', methods=['POST'])
def export_revision():
    """Экспорт отчёта по ревизии"""
    if 'user_id' not in session:
        return jsonify({'error': 'Не авторизован'}), 401

    data = request.json or {}
    fmt = data.get('format', 'excel')
    revision_id = data.get('revision_id')

    try:
        db = get_db_connection()
        cursor = db.cursor()

        if revision_id:
            cursor.execute('''
                SELECT rt.*, p.name as product_name, p.barcode_main
                FROM revision_transactions rt
                LEFT JOIN products_1c p ON rt.product_id = p.id
                WHERE rt.revision_id = ?
                ORDER BY rt.created_at
            ''', (revision_id,))
        else:
            cursor.execute('''
                SELECT rt.*, p.name as product_name, p.barcode_main
                FROM revision_transactions rt
                LEFT JOIN products_1c p ON rt.product_id = p.id
                ORDER BY rt.created_at DESC
                LIMIT 1000
            ''')

        rows = cursor.fetchall()
        db.close()

        action_labels = {
            'sold': 'Продажа',
            'sold_discount': 'Продажа со скидкой',
            'sold_promo': 'Продажа по акции',
            'written_off_expired': 'Списание (просрочка)',
            'written_off_damaged': 'Списание (повреждение)',
            'taken_personal': 'Личное использование',
            'added': 'Добавление',
            'updated': 'Обновление'
        }

        headers = ['ID', 'Товар', 'Штрих-код', 'Действие', 'Количество', 'Цена', 'Скидка', 'Дата']
        export_rows = []

        for row in rows:
            export_rows.append([
                row['id'],
                row['product_name'] or '',
                row['barcode_main'] or '',
                action_labels.get(row['action'], row['action']),
                row['quantity'] or 0,
                row['price_with_discount'] or 0,
                row['discount_percent'] or 0,
                row['created_at'] or ''
            ])

        report_data = {
            'title': f'Отчёт по ревизии #{revision_id}' if revision_id else 'Отчёт по операциям',
            'headers': headers,
            'rows': export_rows
        }

        if fmt == 'excel':
            filename = f'revision_{revision_id or "all"}.xlsx'
            filepath = generate_excel_report(report_data, filename)
            if filepath:
                return send_file(filepath, as_attachment=True, download_name=filename,
                               mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            return jsonify({'error': 'Ошибка генерации Excel'}), 500

        elif fmt == 'pdf':
            filename = f'revision_{revision_id or "all"}.pdf'
            filepath = generate_pdf_report(report_data, filename)
            if filepath:
                return send_file(filepath, as_attachment=True, download_name=filename,
                               mimetype='application/pdf')
            return jsonify({'error': 'Ошибка генерации PDF'}), 500

        return jsonify({'error': 'Неподдерживаемый формат'}), 400

    except Exception as e:
        logger.error(f"Error exporting revision: {e}")
        return jsonify({'error': str(e)}), 500
