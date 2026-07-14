# -*- coding: utf-8 -*-
"""
web_api.py — API роуты для графика, задач, файлов
"""
from flask import Blueprint, request, jsonify, session
from datetime import datetime
import json
import os
from web_config import logger, get_db_connection, audit_log, send_telegram_message, DATA_DIR, UPLOADS_DIR
import recurring_schedule

api_bp = Blueprint('api', __name__)


@api_bp.route('/api/schedule')
def api_get_schedule():
    """Get schedule for month"""
    logger.debug(f"=== API: GET /api/schedule ===")
    logger.debug(f"  Session: user_id={session.get('user_id')}, username={session.get('username')}")
    logger.debug(f"  Request args: {request.args.to_dict()}")
    
    if 'user_id' not in session:
        logger.warning("Not authorized")
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401

    user_id = request.args.get('user_id', session['user_id'], type=int)
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    role = session.get('role', 'employee')
    
    logger.debug(f"  Params: user_id={user_id}, year={year}, month={month}, role={role}")

    db = get_db_connection()
    cursor = db.cursor()

    if role == 'admin':
        logger.debug("  Admin mode: fetching all users")
        cursor.execute('''
            SELECT s.*, u.full_name
            FROM work_schedule s
            JOIN users u ON s.user_id = u.id
            WHERE s.year = ? AND s.month = ?
            ORDER BY s.day, u.full_name
        ''', (year, month))
    else:
        logger.debug(f"  Employee mode: fetching for user_id={user_id}")
        cursor.execute('''
            SELECT * FROM work_schedule
            WHERE user_id = ? AND year = ? AND month = ?
            ORDER BY day
        ''', (user_id, year, month))

    schedule = [dict(row) for row in cursor.fetchall()]
    logger.debug(f"  Found {len(schedule)} schedule entries")

    # Get recurring tasks
    try:
        recurring = recurring_schedule.get_recurring_for_month(year, month)
        logger.debug(f"  Recurring tasks: {len(recurring) if recurring else 0}")
    except Exception as e:
        logger.warning(f"  Recurring schedule error: {e}")
        recurring = []

    return jsonify({'status': 'success', 'schedule': schedule, 'recurring': recurring})


@api_bp.route('/api/schedule/update', methods=['POST'])
def api_update_schedule():
    """Update schedule"""
    logger.debug(f"=== API: POST /api/schedule/update ===")
    logger.debug(f"  Session: user_id={session.get('user_id')}, role={session.get('role')}")
    
    if 'user_id' not in session:
        logger.warning("Not authorized")
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401

    data = request.json or {}
    user_id = data.get('user_id', session['user_id'])
    year = data.get('year')
    month = data.get('month')
    day = data.get('day')
    task_ids = data.get('task_ids', [])
    notes = data.get('notes', '')
    
    logger.debug(f"  Data: user_id={user_id}, date={year}-{month}-{day}, task_ids={task_ids}, notes={notes}")

    if not all([year, month, day]):
        logger.warning("Date required")
        return jsonify({'status': 'error', 'message': 'Date required'}), 400

    try:
        if isinstance(task_ids, str):
            task_ids = json.loads(task_ids) if task_ids.strip() else []
        if not isinstance(task_ids, list):
            task_ids = []
        task_ids = [int(x) for x in task_ids if x is not None and str(x).strip() != '']
    except Exception as e:
        logger.warning(f"  Task IDs parse error: {e}")
        task_ids = []

    try:
        db = get_db_connection()
        cursor = db.cursor()
        logger.debug(f"  Executing INSERT OR REPLACE INTO work_schedule...")
        cursor.execute('''
            INSERT OR REPLACE INTO work_schedule (user_id, year, month, day, task_ids, notes)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (int(user_id), int(year), int(month), int(day), json.dumps(task_ids), notes))
        db.commit()
        logger.debug(f"  Schedule updated successfully")

        # Emit socket event
        from web_server import socketio
        socketio.emit('schedule_updated', {
            'user_id': int(user_id),
            'year': int(year),
            'month': int(month),
            'day': int(day)
        })
        logger.debug(f"  Socket event emitted: schedule_updated")

        logger.info(f"Schedule updated: user_id={user_id}, date={year}-{month}-{day}")
        return jsonify({'status': 'success'})
    except Exception as e:
        logger.exception(f"Schedule update error: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


@api_bp.route('/api/tasks', methods=['GET', 'POST'])
def api_tasks():
    """Get all tasks or create new task"""
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401

    db = get_db_connection()
    cursor = db.cursor()

    # GET - получить все задачи
    if request.method == 'GET':
        cursor.execute('SELECT id, name, color, is_active FROM tasks ORDER BY name')
        tasks = [dict(row) for row in cursor.fetchall()]
        return jsonify({'status': 'success', 'tasks': tasks})

    # POST - создать новую задачу
    data = request.json or {}
    name = data.get('name', '').strip()
    color = data.get('color', '#3498db')

    if not name:
        return jsonify({'status': 'error', 'message': 'Название задачи обязательно'}), 400

    # Проверяем существование задачи с таким названием
    cursor.execute('SELECT id FROM tasks WHERE name = ?', (name,))
    if cursor.fetchone():
        return jsonify({'status': 'error', 'message': 'Задача с таким названием уже существует'}), 409

    cursor.execute('''
        INSERT INTO tasks (name, color, created_by, is_active)
        VALUES (?, ?, ?, 1)
    ''', (name, color, session['user_id']))

    db.commit()
    task_id = cursor.lastrowid

    logger.info(f"Task created: id={task_id}, name={name}, created_by={session['username']}")

    return jsonify({
        'status': 'success',
        'message': 'Задача создана',
        'task_id': task_id
    })


@api_bp.route('/api/tasks/<int:task_id>', methods=['GET', 'PUT', 'DELETE'])
def api_task_by_id(task_id):
    """Get, update or delete a specific task"""
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401

    db = get_db_connection()
    cursor = db.cursor()

    # GET - получить информацию о задаче
    if request.method == 'GET':
        cursor.execute('SELECT id, name, color, is_active, created_by, created_at FROM tasks WHERE id = ?', (task_id,))
        task = cursor.fetchone()

        if not task:
            return jsonify({'status': 'error', 'message': 'Задача не найдена'}), 404

        return jsonify({'status': 'success', 'task': dict(task)})

    # PUT - обновить задачу
    if request.method == 'PUT':
        data = request.json or {}
        name = data.get('name', '').strip()
        color = data.get('color')
        is_active = data.get('is_active')

        # Проверяем существование задачи
        cursor.execute('SELECT * FROM tasks WHERE id = ?', (task_id,))
        task = cursor.fetchone()

        if not task:
            return jsonify({'status': 'error', 'message': 'Задача не найдена'}), 404

        # Обновляем поля
        updates = []
        values = []
        if name:
            updates.append('name = ?')
            values.append(name)
        if color:
            updates.append('color = ?')
            values.append(color)
        if is_active is not None:
            updates.append('is_active = ?')
            values.append(1 if is_active else 0)

        if updates:
            values.append(task_id)
            cursor.execute(f'''
                UPDATE tasks SET {', '.join(updates)} WHERE id = ?
            ''', values)
            db.commit()
            logger.info(f"Task updated: id={task_id}, updates={updates}")

        return jsonify({'status': 'success', 'message': 'Задача обновлена'})

    # DELETE - удалить задачу
    if request.method == 'DELETE':
        # Проверяем существование задачи
        cursor.execute('SELECT * FROM tasks WHERE id = ?', (task_id,))
        task = cursor.fetchone()

        if not task:
            return jsonify({'status': 'error', 'message': 'Задача не найдена'}), 404

        # Нельзя удалить стандартные задачи
        task_name = task['name']
        standard_tasks = ['Протирка полок', 'Уборка влажная', 'Проверка ценников', 
                         'Сроки годности, Акции', 'Ревизия', 'Смена физическая', 
                         'Отпуск', 'Больничный', 'Выходной']
        
        if task_name in standard_tasks:
            return jsonify({'status': 'error', 'message': 'Нельзя удалить стандартную задачу'}), 403

        # Проверяем есть ли эта задача в расписании
        cursor.execute('SELECT id FROM work_schedule WHERE task_ids LIKE ?', (f'%{task_id}%',))
        if cursor.fetchone():
            # Не удаляем, а деактивируем
            cursor.execute('UPDATE tasks SET is_active = 0 WHERE id = ?', (task_id,))
            db.commit()
            logger.info(f"Task deactivated (used in schedule): id={task_id}")
            return jsonify({'status': 'success', 'message': 'Задача деактивирована (используется в расписании)'})

        cursor.execute('DELETE FROM tasks WHERE id = ?', (task_id,))
        db.commit()

        logger.info(f"Task deleted: id={task_id}, name={task['name']}")
        return jsonify({'status': 'success', 'message': 'Задача удалена'})


@api_bp.route('/api/tasks/add', methods=['POST'])
def api_tasks_add():
    """Create new task (legacy endpoint)"""
    return api_tasks()


@api_bp.route('/api/reminders')
def api_reminders():
    """Get reminders for today and upcoming days"""
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401
    
    try:
        import calendar as cal_mod
        now = datetime.now()
        t_year, t_month, today_d = now.year, now.month, now.day
        db = get_db_connection()
        cursor = db.cursor()
        
        # Today's tasks
        cursor.execute('''
            SELECT s.task_ids FROM work_schedule s
            WHERE s.user_id = ? AND s.year = ? AND s.month = ? AND s.day = ?
        ''', (session['user_id'], t_year, t_month, today_d))
        
        task_names = []
        for row in cursor.fetchall():
            try:
                task_ids = json.loads(row['task_ids']) if row['task_ids'] else []
                for tid in task_ids:
                    cursor.execute('SELECT name FROM tasks WHERE id = ?', (tid,))
                    t = cursor.fetchone()
                    if t:
                        task_names.append(t['name'])
            except:
                pass
        
        today_summary = ', '.join(task_names) if task_names else '(no tasks)'
        
        # Upcoming reminders
        reminders = list(recurring_schedule.get_upcoming_reminders(days_ahead=3))
        
        # Check tomorrow — ищем смену с "Смена физическая" (task_id=6)
        tomorrow_d = now.day + 1
        days_in_month = cal_mod.monthrange(t_year, t_month)[1]

        if tomorrow_d <= days_in_month:
            # Для админа проверяем есть ли хоть кто-то из сотрудников на завтра
            check_user_id = session['user_id'] if session['user_id'] != 1 else 2
            cursor.execute(
                '''SELECT 1 FROM work_schedule
                   WHERE user_id = ? AND year = ? AND month = ? AND day = ?
                   AND (task_ids LIKE '%6%' OR task_ids LIKE '%["6"]%')
                   LIMIT 1''',
                (check_user_id, t_year, t_month, tomorrow_d)
            )
            if not cursor.fetchone():
                reminders.append({
                    'message': f'Напоминание: {tomorrow_d}.{t_month}.{t_year} - На завтра нет графика. Создайте задачи заранее.'
                })
        else:
            next_month = 1 if t_month == 12 else t_month + 1
            next_year = t_year + 1 if t_month == 12 else t_year
            check_user_id = session['user_id'] if session['user_id'] != 1 else 2
            cursor.execute(
                '''SELECT 1 FROM work_schedule
                   WHERE user_id = ? AND year = ? AND month = ? AND day = 1
                   AND (task_ids LIKE '%6%' OR task_ids LIKE '%["6"]%')
                   LIMIT 1''',
                (check_user_id, next_year, next_month)
            )
            if not cursor.fetchone():
                reminders.append({
                    'message': f'Напоминание: 1.{next_month}.{next_year} - На следующий месяц нет графика. Создайте задачи заранее.'
                })
        
        return jsonify({
            'status': 'success',
            'today_summary': today_summary,
            'today_date': now.strftime('%d.%m.%Y'),
            'reminders': reminders
        })
    except Exception as e:
        logger.exception(f"Reminders error: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


@api_bp.route('/api/files')
def api_get_files():
    """Get files list"""
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401

    db = get_db_connection()
    cursor = db.cursor()
    cursor.execute('''
        SELECT id, filename, filepath, file_type, file_size, uploaded_at
        FROM files
        ORDER BY uploaded_at DESC
    ''')
    files = [dict(row) for row in cursor.fetchall()]

    return jsonify({'status': 'success', 'files': files})


@api_bp.route('/api/files/upload', methods=['POST'])
def api_upload_file():
    """Upload file"""
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401

    if 'file' not in request.files:
        return jsonify({'status': 'error', 'message': 'No file'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'status': 'error', 'message': 'Empty filename'}), 400

    # Сохраняем файл с уникальным именем
    import uuid
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    unique_id = str(uuid.uuid4())[:8]
    name, ext = os.path.splitext(file.filename)
    safe_name = f"{timestamp}_{unique_id}_{name}{ext}"
    filepath = os.path.join(UPLOADS_DIR, safe_name)
    
    # Создаём директорию если не существует
    os.makedirs(UPLOADS_DIR, exist_ok=True)
    file.save(filepath)

    # Записываем в БД
    db = get_db_connection()
    cursor = db.cursor()
    cursor.execute('''
        INSERT INTO files (filename, filepath, user_id, year, month, day, file_type, file_size, uploaded_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (file.filename, filepath, session['user_id'], 
          datetime.now().year, datetime.now().month, datetime.now().day,
          file.content_type, os.path.getsize(filepath), datetime.now()))

    db.commit()
    file_id = cursor.lastrowid

    logger.info(f"File uploaded: id={file_id}, filename={file.filename}")

    return jsonify({
        'status': 'success',
        'message': 'Файл загружен',
        'file_id': file_id
    })


@api_bp.route('/api/files/<int:file_id>', methods=['GET', 'DELETE'])
def api_get_or_delete_file(file_id):
    """Get file info or delete file"""
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401

    db = get_db_connection()
    cursor = db.cursor()

    # GET - получить информацию о файле
    if request.method == 'GET':
        cursor.execute('SELECT * FROM files WHERE id = ?', (file_id,))
        file = cursor.fetchone()

        if not file:
            logger.warning(f"File {file_id} not found")
            return jsonify({'status': 'error', 'message': 'File not found'}), 404

        file = dict(file)
        logger.debug(f"Get file info: id={file_id}, filename={file['filename']}")
        return jsonify({'status': 'success', 'file': file})

    # DELETE - удалить файл
    logger.debug(f"=== API: DELETE /api/files/{file_id} ===")

    cursor.execute('SELECT * FROM files WHERE id = ?', (file_id,))
    file = cursor.fetchone()

    if not file:
        logger.warning(f"File {file_id} not found")
        return jsonify({'status': 'error', 'message': 'File not found'}), 404

    file = dict(file)
    logger.debug(f"File to delete: {file['filename']}, filepath: {file['filepath']}")

    # Удаляем физический файл
    try:
        if os.path.exists(file['filepath']):
            os.remove(file['filepath'])
            logger.debug(f"Physical file deleted: {file['filepath']}")
    except Exception as e:
        logger.warning(f"Failed to delete physical file: {e}")

    # Удаляем запись из БД
    cursor.execute('DELETE FROM files WHERE id = ?', (file_id,))
    db.commit()

    logger.info(f"File deleted: id={file_id}, filename={file['filename']}")

    return jsonify({'status': 'success', 'message': 'Файл удалён'})


@api_bp.route('/api/files/<int:file_id>/view', methods=['GET'])
def api_view_file(file_id):
    """Просмотр файла"""
    logger.debug(f"=== API: GET /api/files/{file_id}/view ===")
    
    if 'user_id' not in session:
        logger.warning("Not authorized")
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401

    db = get_db_connection()
    cursor = db.cursor()
    
    # Получаем информацию о файле
    cursor.execute('SELECT * FROM files WHERE id = ?', (file_id,))
    file = cursor.fetchone()
    
    if not file:
        logger.warning(f"File {file_id} not found")
        return jsonify({'status': 'error', 'message': 'File not found'}), 404
    
    file = dict(file)
    logger.debug(f"File to view: {file['filename']}, filepath: {file['filepath']}")
    
    # Проверяем существование файла
    if not os.path.exists(file['filepath']):
        logger.warning(f"File not found on disk: {file['filepath']}")
        return jsonify({'status': 'error', 'message': 'Файл не найден'}), 404
    
    # Определяем MIME тип
    import mimetypes
    mime_type, _ = mimetypes.guess_type(file['filename'])
    if not mime_type:
        mime_type = 'application/octet-stream'
    
    # Отправляем файл
    from flask import send_file
    logger.info(f"Sending file: id={file_id}, filename={file['filename']}")
    return send_file(file['filepath'], mimetype=mime_type)


@api_bp.route('/api/files/<int:file_id>/download', methods=['GET'])
def api_download_file(file_id):
    """Скачать файл"""
    logger.debug(f"=== API: GET /api/files/{file_id}/download ===")
    
    if 'user_id' not in session:
        logger.warning("Not authorized")
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401

    db = get_db_connection()
    cursor = db.cursor()
    
    cursor.execute('SELECT * FROM files WHERE id = ?', (file_id,))
    file = cursor.fetchone()
    
    if not file:
        logger.warning(f"File {file_id} not found")
        return jsonify({'status': 'error', 'message': 'File not found'}), 404
    
    file = dict(file)
    
    if not os.path.exists(file['filepath']):
        logger.warning(f"File not found on disk: {file['filepath']}")
        return jsonify({'status': 'error', 'message': 'Файл не найден'}), 404
    
    from flask import send_file
    logger.info(f"Sending file for download: id={file_id}, filename={file['filename']}")
    return send_file(file['filepath'], as_attachment=True, download_name=file['filename'])


@api_bp.route('/api/users', methods=['GET', 'POST'])
def api_users():
    """Get all users or create new user"""
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401

    db = get_db_connection()
    cursor = db.cursor()

    # GET - получить всех пользователей
    if request.method == 'GET':
        cursor.execute('SELECT id, username, full_name, role FROM users ORDER BY username')
        users = [dict(row) for row in cursor.fetchall()]
        return jsonify({'status': 'success', 'users': users})

    # POST - создать нового пользователя (только админ)
    if session.get('role') != 'admin':
        return jsonify({'status': 'error', 'message': 'Требуется роль администратора'}), 403

    data = request.json or {}
    username = data.get('username', '').strip()
    password = data.get('password', '')
    full_name = data.get('full_name', '').strip()
    role = data.get('role', 'employee')

    if not username:
        return jsonify({'status': 'error', 'message': 'Имя пользователя обязательно'}), 400

    if not password:
        return jsonify({'status': 'error', 'message': 'Пароль обязателен'}), 400

    # Проверяем существование пользователя
    cursor.execute('SELECT id FROM users WHERE username = ?', (username,))
    if cursor.fetchone():
        return jsonify({'status': 'error', 'message': 'Пользователь с таким именем уже существует'}), 409

    import hashlib
    password_hash = hashlib.sha256(password.encode()).hexdigest()

    cursor.execute('''
        INSERT INTO users (username, password_hash, role, full_name)
        VALUES (?, ?, ?, ?)
    ''', (username, password_hash, role, full_name))

    db.commit()
    user_id = cursor.lastrowid

    logger.info(f"User created: id={user_id}, username={username}, role={role}, created_by={session['username']}")

    return jsonify({
        'status': 'success',
        'message': 'Пользователь создан',
        'user_id': user_id
    })


@api_bp.route('/api/users/<int:user_id>', methods=['GET', 'PUT', 'DELETE'])
def api_user_by_id(user_id):
    """Get, update or delete a specific user"""
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401

    # Только админ может управлять пользователями
    if session.get('role') != 'admin':
        return jsonify({'status': 'error', 'message': 'Требуется роль администратора'}), 403

    db = get_db_connection()
    cursor = db.cursor()

    # GET - получить информацию о пользователе
    if request.method == 'GET':
        cursor.execute('SELECT id, username, full_name, role, created_at FROM users WHERE id = ?', (user_id,))
        user = cursor.fetchone()

        if not user:
            return jsonify({'status': 'error', 'message': 'Пользователь не найден'}), 404

        return jsonify({'status': 'success', 'user': dict(user)})

    # PUT - обновить пользователя
    if request.method == 'PUT':
        data = request.json or {}
        full_name = data.get('full_name', '').strip()
        role = data.get('role')

        # Проверяем существование пользователя
        cursor.execute('SELECT * FROM users WHERE id = ?', (user_id,))
        user = cursor.fetchone()

        if not user:
            return jsonify({'status': 'error', 'message': 'Пользователь не найден'}), 404

        # Нельзя редактировать самого себя (роль)
        if user_id == session['user_id'] and role:
            return jsonify({'status': 'error', 'message': 'Нельзя изменить свою собственную роль'}), 400

        # Обновляем поля
        updates = []
        values = []
        if full_name:
            updates.append('full_name = ?')
            values.append(full_name)
        if role and role in ['admin', 'employee']:
            updates.append('role = ?')
            values.append(role)

        if updates:
            values.append(user_id)
            cursor.execute(f'''
                UPDATE users SET {', '.join(updates)} WHERE id = ?
            ''', values)
            db.commit()
            logger.info(f"User updated: id={user_id}, updates={updates}")

        return jsonify({'status': 'success', 'message': 'Пользователь обновлён'})

    # DELETE - удалить пользователя
    if request.method == 'DELETE':
        # Нельзя удалить самого себя
        if user_id == session['user_id']:
            return jsonify({'status': 'error', 'message': 'Нельзя удалить самого себя'}), 400

        # Проверяем существование пользователя
        cursor.execute('SELECT * FROM users WHERE id = ?', (user_id,))
        user = cursor.fetchone()

        if not user:
            return jsonify({'status': 'error', 'message': 'Пользователь не найден'}), 404

        # Нельзя удалить последнего админа
        if user['role'] == 'admin':
            cursor.execute('SELECT COUNT(*) as count FROM users WHERE role = ? AND id != ?', ('admin', user_id))
            admin_count = cursor.fetchone()['count']
            if admin_count < 1:
                return jsonify({'status': 'error', 'message': 'Нельзя удалить последнего админа'}), 403

        # Проверяем есть ли записи в расписании
        cursor.execute('SELECT COUNT(*) as count FROM work_schedule WHERE user_id = ?', (user_id,))
        schedule_count = cursor.fetchone()['count']
        if schedule_count > 0:
            return jsonify({'status': 'error', 'message': 'Нельзя удалить пользователя с записями в расписании. Сначала удалите расписание.'}), 403

        cursor.execute('DELETE FROM users WHERE id = ?', (user_id,))
        db.commit()

        logger.info(f"User deleted: id={user_id}, username={user['username']}")
        return jsonify({'status': 'success', 'message': 'Пользователь удалён'})


@api_bp.route('/api/users/<int:user_id>/password', methods=['PUT'])
def api_user_password(user_id):
    """Update user password"""
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401

    # Только админ или сам пользователь может менять пароль
    if session.get('role') != 'admin' and session['user_id'] != user_id:
        return jsonify({'status': 'error', 'message': 'Требуется роль администратора'}), 403

    data = request.json or {}
    new_password = data.get('new_password', '')

    if not new_password:
        return jsonify({'status': 'error', 'message': 'Новый пароль обязателен'}), 400

    if len(new_password) < 4:
        return jsonify({'status': 'error', 'message': 'Пароль должен быть не менее 4 символов'}), 400

    import hashlib
    password_hash = hashlib.sha256(new_password.encode()).hexdigest()

    db = get_db_connection()
    cursor = db.cursor()

    # Проверяем существование пользователя
    cursor.execute('SELECT * FROM users WHERE id = ?', (user_id,))
    user = cursor.fetchone()

    if not user:
        return jsonify({'status': 'error', 'message': 'Пользователь не найден'}), 404

    cursor.execute('UPDATE users SET password_hash = ? WHERE id = ?', (password_hash, user_id))
    db.commit()

    logger.info(f"User password updated: id={user_id}, username={user['username']}")

    return jsonify({'status': 'success', 'message': 'Пароль обновлён'})


@api_bp.route('/api/colleagues')
def api_get_colleagues():
    """Get colleagues list"""
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401

    db = get_db_connection()
    cursor = db.cursor()
    cursor.execute('''
        SELECT id, username, full_name, role
        FROM users
        WHERE id != ?
        ORDER BY full_name
    ''', (session['user_id'],))
    colleagues = [dict(row) for row in cursor.fetchall()]

    return jsonify({'status': 'success', 'colleagues': colleagues})


@api_bp.route('/api/colleague-tasks', methods=['GET', 'POST'])
def api_colleague_tasks():
    """Get colleague tasks for schedule or create new task"""
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401

    # POST - создать новую задачу для коллеги
    if request.method == 'POST':
        data = request.json or {}
        user_id = data.get('user_id')
        year = data.get('year')
        month = data.get('month')
        day = data.get('day')
        task_ids = data.get('task_ids', [])
        notes = data.get('notes', '')

        if not all([user_id, year, month, day]):
            return jsonify({'status': 'error', 'message': 'user_id, year, month, day обязательны'}), 400

        db = get_db_connection()
        cursor = db.cursor()

        # Проверяем существование пользователя
        cursor.execute('SELECT * FROM users WHERE id = ?', (user_id,))
        user = cursor.fetchone()

        if not user:
            return jsonify({'status': 'error', 'message': 'Пользователь не найден'}), 404

        try:
            if isinstance(task_ids, str):
                task_ids = json.loads(task_ids) if task_ids.strip() else []
            if not isinstance(task_ids, list):
                task_ids = []
            task_ids = [int(x) for x in task_ids if x is not None and str(x).strip() != '']
        except Exception as e:
            logger.warning(f"Task IDs parse error: {e}")
            task_ids = []

        cursor.execute('''
            INSERT OR REPLACE INTO work_schedule (user_id, year, month, day, task_ids, notes)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (int(user_id), int(year), int(month), int(day), json.dumps(task_ids), notes))

        db.commit()

        logger.info(f"Colleague task created: user_id={user_id}, date={year}-{month}-{day}, created_by={session['username']}")

        return jsonify({
            'status': 'success',
            'message': 'Задача создана'
        })

    # GET - получить задачи
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    role = session.get('role', 'employee')

    db = get_db_connection()
    cursor = db.cursor()

    if role == 'admin':
        # Админ видит все задачи всех сотрудников
        cursor.execute('''
            SELECT s.*, u.full_name, u.username
            FROM work_schedule s
            JOIN users u ON s.user_id = u.id
            WHERE s.year = ? AND s.month = ?
            ORDER BY s.day, u.full_name
        ''', (year, month))
    else:
        # Сотрудник видит только свои задачи
        cursor.execute('''
            SELECT s.*, u.full_name, u.username
            FROM work_schedule s
            JOIN users u ON s.user_id = u.id
            WHERE s.year = ? AND s.month = ? AND s.user_id = ?
            ORDER BY s.day
        ''', (year, month, session['user_id']))

    tasks = [dict(row) for row in cursor.fetchall()]

    # Получаем названия задач
    cursor.execute('SELECT id, name, color FROM tasks WHERE is_active = 1')
    all_tasks = {row['id']: {'name': row['name'], 'color': row['color']} for row in cursor.fetchall()}

    # Добавляем расшифровку task_ids
    for task in tasks:
        try:
            task_ids = json.loads(task.get('task_ids', '[]')) if task.get('task_ids') else []
            task['task_names'] = [all_tasks.get(int(tid), {}).get('name', 'Неизвестно') for tid in task_ids if tid]
        except:
            task['task_names'] = []

    logger.debug(f"Colleague tasks: found {len(tasks)} entries for {year}-{month}")

    return jsonify({'status': 'success', 'tasks': tasks})


@api_bp.route('/api/colleague-tasks/<int:task_id>/complete', methods=['PATCH'])
def api_colleague_task_complete(task_id):
    """Mark colleague task as complete"""
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401

    data = request.json or {}
    completed = data.get('completed', False)

    db = get_db_connection()
    cursor = db.cursor()

    # Проверяем существование записи
    cursor.execute('SELECT * FROM work_schedule WHERE id = ?', (task_id,))
    task = cursor.fetchone()

    if not task:
        return jsonify({'status': 'error', 'message': 'Запись не найдена'}), 404

    # Для простоты просто логируем завершение
    logger.info(f"Colleague task {task_id} marked as {'completed' if completed else 'incomplete'}")

    return jsonify({
        'status': 'success',
        'message': f'Задача {"завершена" if completed else "возвращена в работу"}'
    })


@api_bp.route('/api/colleague-tasks/<int:task_id>/send-telegram', methods=['POST'])
def api_colleague_task_send_telegram(task_id):
    """Send colleague task to Telegram"""
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401

    db = get_db_connection()
    cursor = db.cursor()

    # Получаем запись
    cursor.execute('''
        SELECT s.*, u.full_name, u.username
        FROM work_schedule s
        JOIN users u ON s.user_id = u.id
        WHERE s.id = ?
    ''', (task_id,))
    task = cursor.fetchone()

    if not task:
        return jsonify({'status': 'error', 'message': 'Запись не найдена'}), 404

    task = dict(task)

    # Получаем названия задач
    try:
        task_ids = json.loads(task.get('task_ids', '[]')) if task.get('task_ids') else []
        cursor.execute('SELECT id, name FROM tasks WHERE id IN ({})'.format(','.join('?' * len(task_ids))), task_ids)
        task_names = [row['name'] for row in cursor.fetchall()]
    except:
        task_names = []

    # Формируем сообщение
    message = f"📋 <b>Задача на {task['day']}.{task['month']}.{task['year']}</b>\n"
    message += f"👤 <b>Сотрудник:</b> {task.get('full_name', task.get('username', 'Неизвестно'))}\n"
    message += f"✅ <b>Задачи:</b> {', '.join(task_names) if task_names else 'Нет задач'}"

    if task.get('notes'):
        message += f"\n📝 <b>Заметки:</b> {task['notes']}"

    # Отправляем в Telegram
    try:
        config_path = os.path.join(DATA_DIR, 'telegram_config.json')
        with open(config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
        token = config.get('token')
        chat_ids = config.get('chat_ids', [])

        if token and chat_ids:
            sent, total = send_telegram_message(chat_ids, message, token)
            logger.info(f"Telegram sent: {sent}/{total}")

            return jsonify({
                'status': 'success',
                'message': f'Отправлено в Telegram: {sent}/{total}'
            })
        else:
            return jsonify({'status': 'error', 'message': 'Telegram не настроен'}), 400
    except Exception as e:
        logger.exception(f"Telegram send error: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


@api_bp.route('/api/colleague-tasks/<int:task_id>/thanks', methods=['POST'])
def api_colleague_task_thanks(task_id):
    """Send thanks for colleague task"""
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401

    data = request.json or {}
    thanks_message = data.get('message', 'Спасибо за работу!')

    db = get_db_connection()
    cursor = db.cursor()

    # Получаем запись
    cursor.execute('''
        SELECT s.*, u.full_name, u.username
        FROM work_schedule s
        JOIN users u ON s.user_id = u.id
        WHERE s.id = ?
    ''', (task_id,))
    task = cursor.fetchone()

    if not task:
        return jsonify({'status': 'error', 'message': 'Запись не найдена'}), 404

    task = dict(task)

    # Логируем благодарность
    cursor.execute('''
        INSERT INTO audit_log (user_id, action, details, created_at)
        VALUES (?, ?, ?, ?)
    ''', (task['user_id'], 'thanks', json.dumps({
        'from_user_id': session['user_id'],
        'from_username': session['username'],
        'message': thanks_message,
        'schedule_id': task_id
    }), datetime.now()))

    db.commit()

    logger.info(f"Thanks sent: from={session['username']}, to={task.get('username')}, message={thanks_message}")

    return jsonify({
        'status': 'success',
        'message': 'Благодарность отправлена'
    })


@api_bp.route('/chat/topics')
def api_get_chat_topics():
    """Get chat topics"""
    db = get_db_connection()
    cursor = db.cursor()
    cursor.execute('SELECT id, title FROM chat_topics ORDER BY id')
    topics = [dict(row) for row in cursor.fetchall()]
    
    return jsonify({'status': 'success', 'topics': topics})


@api_bp.route('/chat/topics', methods=['POST'])
def api_create_chat_topic():
    """Create chat topic"""
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401
    
    data = request.json or {}
    title = data.get('title', '').strip()
    
    if not title:
        return jsonify({'status': 'error', 'message': 'Title required'}), 400
    
    db = get_db_connection()
    cursor = db.cursor()
    
    cursor.execute('''
        INSERT INTO chat_topics (title, created_by, created_at)
        VALUES (?, ?, ?)
    ''', (title, session['user_id'], datetime.now()))
    
    db.commit()
    topic_id = cursor.lastrowid
    
    logger.info(f"Chat topic created: id={topic_id}, title={title}")
    
    return jsonify({
        'status': 'success',
        'message': 'Тема создана',
        'topic_id': topic_id
    })


@api_bp.route('/chat/topics/<int:topic_id>', methods=['PUT'])
def api_update_chat_topic(topic_id):
    """Update chat topic"""
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401
    
    data = request.json or {}
    title = data.get('title', '').strip()
    
    if not title:
        return jsonify({'status': 'error', 'message': 'Title required'}), 400
    
    db = get_db_connection()
    cursor = db.cursor()
    
    cursor.execute('''
        UPDATE chat_topics SET title = ? WHERE id = ?
    ''', (title, topic_id))
    
    db.commit()
    
    logger.info(f"Chat topic updated: id={topic_id}, title={title}")
    
    return jsonify({'status': 'success', 'message': 'Тема обновлена'})


@api_bp.route('/chat/topics/<int:topic_id>', methods=['DELETE'])
def api_delete_chat_topic(topic_id):
    """Delete chat topic"""
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401
    
    db = get_db_connection()
    cursor = db.cursor()
    
    # Нельзя удалить общий чат (id=1)
    if topic_id == 1:
        return jsonify({'status': 'error', 'message': 'Нельзя удалить общий чат'}), 400
    
    cursor.execute('''
        DELETE FROM chat_topics WHERE id = ?
    ''', (topic_id,))
    
    db.commit()
    
    logger.info(f"Chat topic deleted: id={topic_id}")
    
    return jsonify({'status': 'success', 'message': 'Тема удалена'})


@api_bp.route('/chat/send', methods=['POST'])
def api_send_chat_message():
    """Send chat message"""
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401
    
    data = request.json or {}
    message = data.get('message', '').strip()
    topic_id = int(data.get('topic_id', 1))
    attachment_file_id = data.get('attachment_file_id')

    # Разрешаем отправку только файла без текста
    if not message and not attachment_file_id:
        return jsonify({'status': 'error', 'message': 'Message or attachment required'}), 400
    
    db = get_db_connection()
    cursor = db.cursor()
    
    cursor.execute('''
        INSERT INTO chat_messages
        (user_id, username, full_name, message, topic_id, attachment_file_id, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (
        session['user_id'],
        session['username'],
        session.get('full_name', ''),
        message if message else '',
        topic_id,
        attachment_file_id,
        datetime.now()
    ))
    
    db.commit()
    msg_id = cursor.lastrowid

    logger.info(f"Chat message saved: id={msg_id}, user={session['username']}, topic_id={topic_id}")

    # Отправляем в Socket.IO через server-side emit
    try:
        from web_server import socketio
        socketio.emit('chat_message', {
            'id': msg_id,
            'user_id': session['user_id'],
            'username': session['username'],
            'full_name': session.get('full_name', ''),
            'message': message if message else '',
            'topic_id': topic_id,
            'attachment_file_id': attachment_file_id,
            'created_at': datetime.now().isoformat()
        })  # broadcast=True удалён
        logger.debug(f"Socket event emitted: chat_message")
    except Exception as e:
        logger.warning(f"Socket emit warning: {e}")

    # Отправляем в VK группу (если настроено и это тема VK)
    if topic_id == 3:  # VK тема
        logger.info(f"🔵 VK тема обнаружена, пытаемся отправить в VK...")
        try:
            # Пробуем загрузить vk_config напрямую
            config_path = os.path.join(DATA_DIR, 'vk_config.json')
            vk_config = {}
            if os.path.exists(config_path):
                with open(config_path, 'r', encoding='utf-8') as f:
                    vk_config = json.load(f)
                logger.info(f"✅ VK config загружен: group_id={vk_config.get('group_id')}, chat_peer_id={vk_config.get('chat_peer_id')}")
            else:
                logger.warning(f"❌ VK config не найден: {config_path}")

            chat_peer_id = vk_config.get('chat_peer_id')

            if chat_peer_id:
                # Формируем сообщение от имени пользователя
                vk_message = f"[{session.get('full_name', session['username'])}]: {message}" if message else ''

                # Если есть вложение, добавляем ссылку
                if attachment_file_id:
                    vk_message += f" [Файл: {attachment_file_id}]"

                logger.info(f"📤 Отправляем в VK: {vk_message[:100]}...")

                # Используем urllib напрямую для отправки
                token = vk_config.get('service_token')
                api_version = vk_config.get('api_version', '5.131')

                if token:
                    import urllib.parse
                    import ssl
                    import urllib.request

                    url = "https://api.vk.com/method/messages.send"
                    params = {
                        'peer_id': int(chat_peer_id),
                        'message': vk_message,
                        'random_id': int(datetime.now().timestamp() * 1000),
                        'access_token': token,
                        'v': api_version
                    }

                    data = urllib.parse.urlencode(params).encode('utf-8')
                    req = urllib.request.Request(url, data=data, method='POST')
                    req.add_header('Content-Type', 'application/x-www-form-urlencoded')

                    _ssl_context = ssl.create_default_context()
                    _ssl_context.check_hostname = False
                    _ssl_context.verify_mode = ssl.CERT_NONE

                    try:
                        with urllib.request.urlopen(req, timeout=10, context=_ssl_context) as response:
                            result = json.loads(response.read().decode('utf-8'))

                        if 'error' in result:
                            logger.error(f"❌ VK API error: {result['error']}")
                        else:
                            logger.info(f"✅ VK сообщение отправлено успешно! response={result.get('response')}")
                    except Exception as http_error:
                        logger.exception(f"❌ HTTP ошибка при отправке в VK: {http_error}")
                else:
                    logger.warning(f"⚠️ VK токен не найден в конфиге")
            else:
                logger.warning(f"⚠️ chat_peer_id не настроен в VK конфиге")
        except Exception as e:
            logger.exception(f"❌ Ошибка отправки в VK: {e}")

    logger.info(f"Chat message processed: topic_id={topic_id}, user={session['username']}")

    return jsonify({
        'status': 'success',
        'message': 'Сообщение отправлено',
        'id': msg_id
    })


@api_bp.route('/chat/messages')
def api_get_chat_messages():
    """Get chat messages"""
    topic_id = request.args.get('topic_id', 1, type=int)
    limit = request.args.get('limit', 50, type=int)

    db = get_db_connection()
    cursor = db.cursor()
    cursor.execute('''
        SELECT m.id, m.user_id, m.username, m.full_name, m.message, m.attachment_file_id, m.created_at
        FROM chat_messages m
        WHERE m.topic_id = ?
        ORDER BY m.created_at ASC
        LIMIT ?
    ''', (topic_id, limit))
    messages = [dict(row) for row in cursor.fetchall()]

    return jsonify({'status': 'success', 'messages': messages})


# ==========================================
# FILES & TUNNEL API
# ==========================================

@api_bp.route('/files/<int:file_id>')
def api_get_file(file_id):
    """Получить информацию о файле"""
    db = get_db_connection()
    cursor = db.cursor()
    cursor.execute('SELECT * FROM files WHERE id = ?', (file_id,))
    file = cursor.fetchone()
    
    if not file:
        return jsonify({'status': 'error', 'message': 'File not found'}), 404
    
    return jsonify({'status': 'success', 'file': dict(file)})


@api_bp.route('/tunnel-info')
def api_tunnel_info():
    """Информация о туннеле"""
    try:
        tunnel_info_path = os.path.join(DATA_DIR, 'tunnel_info.json')
        if os.path.exists(tunnel_info_path):
            with open(tunnel_info_path, 'r', encoding='utf-8') as f:
                info = json.load(f)
            return jsonify({
                'status': 'ok',
                'tunnel_url': info.get('tunnel_url', ''),
                'password': info.get('password', '')
            })
    except Exception as e:
        logger.warning(f"Tunnel info error: {e}")
    return jsonify({'status': 'error', 'message': 'Tunnel info not available'})


@api_bp.route('/tunnel-status')
def api_tunnel_status():
    """Полный статус туннеля с графиком"""
    import psutil
    
    # Получаем информацию о туннеле
    tunnel_info = {}
    try:
        tunnel_info_path = os.path.join(DATA_DIR, 'tunnel_info.json')
        if os.path.exists(tunnel_info_path):
            with open(tunnel_info_path, 'r', encoding='utf-8') as f:
                tunnel_info = json.load(f)
    except:
        pass
    
    # Считываем логи туннеля
    tunnel_logs = []
    try:
        log_path = os.path.join(LOGS_DIR, 'tunnel.log')
        if os.path.exists(log_path):
            with open(log_path, 'r', encoding='utf-8') as f:
                tunnel_logs = f.readlines()[-50:]  # Последние 50 строк
    except:
        pass
    
    # Статистика системы
    cpu_percent = psutil.cpu_percent(interval=1)
    memory = psutil.virtual_memory()
    net_io = psutil.net_io_counters()
    
    return jsonify({
        'status': 'success',
        'tunnel': {
            'url': tunnel_info.get('tunnel_url', 'N/A'),
            'password': tunnel_info.get('password', 'N/A'),
            'active': bool(tunnel_info.get('tunnel_url'))
        },
        'system': {
            'cpu': cpu_percent,
            'memory': memory.percent,
            'memory_used': round(memory.used / 1024 / 1024, 2),
            'memory_total': round(memory.total / 1024 / 1024, 2)
        },
        'network': {
            'bytes_sent': net_io.bytes_sent,
            'bytes_recv': net_io.bytes_recv,
            'packets_sent': net_io.packets_sent,
            'packets_recv': net_io.packets_recv
        },
        'logs': tunnel_logs
    })


# ==========================================
# TELEGRAM INTEGRATION API
# ==========================================

@api_bp.route('/api/telegram-user-map', methods=['POST'])
def api_telegram_user_map():
    """Map Telegram user to system user"""
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401

    if session.get('role') != 'admin':
        return jsonify({'status': 'error', 'message': 'Требуется роль администратора'}), 403

    data = request.json or {}
    system_user_id = data.get('user_id')
    telegram_username = data.get('telegram_username')
    telegram_id = data.get('telegram_id')

    if not system_user_id:
        return jsonify({'status': 'error', 'message': 'user_id обязателен'}), 400

    db = get_db_connection()
    cursor = db.cursor()

    # Проверяем существование системного пользователя
    cursor.execute('SELECT * FROM users WHERE id = ?', (system_user_id,))
    user = cursor.fetchone()

    if not user:
        return jsonify({'status': 'error', 'message': 'Пользователь не найден'}), 404

    # Создаём или обновляем маппинг
    # Используем audit_log или создаём новую таблицу при необходимости
    # Для простоты сохраним в audit_log
    cursor.execute('''
        INSERT INTO audit_log (user_id, action, details, created_at)
        VALUES (?, ?, ?, ?)
    ''', (system_user_id, 'telegram_map', json.dumps({
        'telegram_username': telegram_username,
        'telegram_id': telegram_id
    }), datetime.now()))

    db.commit()

    logger.info(f"Telegram user mapped: system_user_id={system_user_id}, telegram={telegram_username}")

    return jsonify({
        'status': 'success',
        'message': 'Пользователь Telegram привязан'
    })


@api_bp.route('/api/telegram-fetch-members', methods=['POST'])
def api_telegram_fetch_members():
    """Fetch Telegram chat members"""
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401

    if session.get('role') != 'admin':
        return jsonify({'status': 'error', 'message': 'Требуется роль администратора'}), 403

    data = request.json or {}
    chat_id = data.get('chat_id')

    if not chat_id:
        # Пытаемся получить из конфига
        try:
            config_path = os.path.join(DATA_DIR, 'telegram_config.json')
            with open(config_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
            chat_id = config.get('chat_ids', [])
            if not chat_id:
                return jsonify({'status': 'error', 'message': 'chat_id не указан'}), 400
            chat_id = chat_id[0] if isinstance(chat_id, list) else chat_id
        except:
            return jsonify({'status': 'error', 'message': 'Не удалось получить chat_id из конфига'}), 400

    # Получаем токен бота
    try:
        config_path = os.path.join(DATA_DIR, 'telegram_config.json')
        with open(config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
        token = config.get('token')
    except:
        return jsonify({'status': 'error', 'message': 'Не удалось получить токен бота'}), 400

    # Получаем список участников чата (для супергруппы)
    try:
        import urllib.request
        url = f"https://api.telegram.org/bot{token}/getChatAdministrators?chat_id={chat_id}"
        
        req = urllib.request.Request(url)
        with urllib.request.urlopen(req, timeout=10) as resp:
            result = json.loads(resp.read().decode('utf-8'))
            
            if result.get('ok'):
                members = result.get('result', [])
                logger.info(f"Telegram members fetched: {len(members)}")
                
                return jsonify({
                    'status': 'success',
                    'members': members
                })
            else:
                logger.warning(f"Telegram API error: {result}")
                return jsonify({
                    'status': 'error',
                    'message': result.get('description', 'Unknown error')
                }), 400
    except Exception as e:
        logger.exception(f"Telegram fetch error: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


@api_bp.route('/api/weekly-digest', methods=['GET'])
def api_weekly_digest():
    """Get weekly digest of schedule and tasks"""
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401

    from datetime import timedelta
    
    now = datetime.now()
    week_start = now - timedelta(days=now.weekday())
    week_end = week_start + timedelta(days=6)
    
    db = get_db_connection()
    cursor = db.cursor()
    
    # Получаем задачи за неделю
    cursor.execute('''
        SELECT s.*, u.full_name
        FROM work_schedule s
        JOIN users u ON s.user_id = u.id
        WHERE (s.year = ? AND s.month = ? AND s.day >= ? AND s.day <= ?)
    ''', (now.year, now.month, week_start.day, week_end.day))
    
    schedule = [dict(row) for row in cursor.fetchall()]
    
    # Получаем задачи
    cursor.execute('SELECT id, name, color FROM tasks WHERE is_active = 1')
    tasks = {row['id']: {'name': row['name'], 'color': row['color']} for row in cursor.fetchall()}
    
    # Формируем отчёт
    digest = {
        'period': f'{week_start.strftime("%d.%m.%Y")} - {week_end.strftime("%d.%m.%Y")}',
        'schedule': [],
        'summary': {}
    }
    
    for entry in schedule:
        try:
            task_ids = json.loads(entry.get('task_ids', '[]')) if entry.get('task_ids') else []
            task_names = [tasks.get(int(tid), {}).get('name', 'Неизвестно') for tid in task_ids if tid]
            
            digest['schedule'].append({
                'date': f'{entry["day"]}.{now.month}.{now.year}',
                'user': entry.get('full_name', 'Unknown'),
                'tasks': task_names
            })
        except:
            pass
    
    # Статистика
    digest['summary'] = {
        'total_entries': len(schedule),
        'unique_users': len(set(e.get('user_id') for e in schedule))
    }
    
    return jsonify({
        'status': 'success',
        'digest': digest
    })


# ==========================================
# SCHEDULE TEMPLATE API
# ==========================================

@api_bp.route('/api/schedule/template', methods=['POST'])
def api_schedule_template():
    """Create schedule from template"""
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401

    if session.get('role') != 'admin':
        return jsonify({'status': 'error', 'message': 'Требуется роль администратора'}), 403

    data = request.json or {}
    template_name = data.get('template_name')
    year = data.get('year')
    month = data.get('month')
    days = data.get('days', [])  # Список дней с задачами

    if not all([year, month, days]):
        return jsonify({'status': 'error', 'message': 'year, month, days обязательны'}), 400

    db = get_db_connection()
    cursor = db.cursor()

    created_count = 0

    for day_data in days:
        day = day_data.get('day')
        user_id = day_data.get('user_id')
        task_ids = day_data.get('task_ids', [])
        notes = day_data.get('notes', '')

        if not all([day, user_id]):
            continue

        cursor.execute('''
            INSERT OR REPLACE INTO work_schedule (user_id, year, month, day, task_ids, notes)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (int(user_id), int(year), int(month), int(day), json.dumps(task_ids), notes))

        created_count += 1

    db.commit()

    logger.info(f"Schedule template applied: {created_count} entries for {year}-{month}")

    return jsonify({
        'status': 'success',
        'message': f'Шаблон применён: {created_count} записей',
        'created_count': created_count
    })


# ==========================================
# SYSTEM HEALTH API
# ==========================================

@api_bp.route('/api/health', methods=['GET'])
def api_health():
    """Health check endpoint"""
    try:
        db = get_db_connection()
        cursor = db.cursor()
        cursor.execute('SELECT 1')
        db_ok = True
    except:
        db_ok = False

    return jsonify({
        'status': 'success' if db_ok else 'degraded',
        'database': 'ok' if db_ok else 'error',
        'timestamp': datetime.now().isoformat()
    })


# ==========================================
# ADDITIONAL BARCODES API
# ==========================================

@api_bp.route('/api/barcodes/<int:barcode_id>', methods=['GET', 'PUT'])
def api_barcode_by_id(barcode_id):
    """Get or update specific barcode"""
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401

    db = get_db_connection()
    cursor = db.cursor()

    # GET - получить информацию о штрих-коде
    if request.method == 'GET':
        cursor.execute('SELECT * FROM barcodes WHERE id = ?', (barcode_id,))
        barcode = cursor.fetchone()

        if not barcode:
            return jsonify({'status': 'error', 'message': 'Штрих-код не найден'}), 404

        return jsonify({'status': 'success', 'barcode': dict(barcode)})

    # PUT - обновить штрих-код
    if session.get('role') != 'admin':
        return jsonify({'status': 'error', 'message': 'Требуется роль администратора'}), 403

    data = request.json or {}
    product_name = data.get('product_name')
    factory_barcode = data.get('factory_barcode')
    internal_barcode = data.get('internal_barcode')

    cursor.execute('SELECT * FROM barcodes WHERE id = ?', (barcode_id,))
    barcode = cursor.fetchone()

    if not barcode:
        return jsonify({'status': 'error', 'message': 'Штрих-код не найден'}), 404

    updates = []
    values = []
    if product_name:
        updates.append('product_name = ?')
        values.append(product_name)
    if factory_barcode is not None:
        updates.append('factory_barcode = ?')
        values.append(factory_barcode)
    if internal_barcode is not None:
        updates.append('internal_barcode = ?')
        values.append(internal_barcode)

    if updates:
        values.append(barcode_id)
        cursor.execute(f'''
            UPDATE barcodes SET {', '.join(updates)} WHERE id = ?
        ''', values)
        db.commit()
        logger.info(f"Barcode updated: id={barcode_id}")

    return jsonify({'status': 'success', 'message': 'Штрих-код обновлён'})


@api_bp.route('/api/barcodes/import-excel', methods=['POST'])
def api_barcodes_import_excel():
    """Import barcodes from Excel"""
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401

    if session.get('role') != 'admin':
        return jsonify({'status': 'error', 'message': 'Требуется роль администратора'}), 403

    if 'file' not in request.files:
        return jsonify({'status': 'error', 'message': 'Файл не загружен'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'status': 'error', 'message': 'Файл не выбран'}), 400

    try:
        import openpyxl
        import tempfile

        # Сохраняем временный файл
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            file.save(tmp.name)
            tmp_path = tmp.name

        wb = openpyxl.load_workbook(tmp_path, read_only=True)
        ws = wb.active

        db = get_db_connection()
        cursor = db.cursor()

        imported = 0
        for row in ws.iter_rows(min_row=2, values_only=True):  # Пропускаем заголовок
            if not row or not row[0]:
                continue

            code = str(row[0]) if row[0] else ''
            name = str(row[1]) if len(row) > 1 and row[1] else ''
            price = float(row[2]) if len(row) > 2 and row[2] else 0

            cursor.execute('''
                INSERT OR REPLACE INTO barcodes (code, name, price, created_by, created_at)
                VALUES (?, ?, ?, ?, ?)
            ''', (code, name, price, session['user_id'], datetime.now()))

            imported += 1

        db.commit()
        wb.close()
        os.unlink(tmp_path)

        logger.info(f"Barcodes imported from Excel: {imported} items")

        return jsonify({
            'status': 'success',
            'message': f'Импортировано {imported} штрих-кодов',
            'imported': imported
        })

    except Exception as e:
        logger.exception(f"Excel import error: {e}")
        return jsonify({'status': 'error', 'message': f'Ошибка импорта: {str(e)}'}), 500


# ==========================================
# ADMIN BACKUP API
# ==========================================

@api_bp.route('/api/admin/backup', methods=['GET'])
def api_admin_backup():
    """Create database backup"""
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401

    if session.get('role') != 'admin':
        return jsonify({'status': 'error', 'message': 'Требуется роль администратора'}), 403

    import shutil
    from datetime import datetime

    backup_dir = os.path.join(DATA_DIR, 'backups')
    os.makedirs(backup_dir, exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_path = os.path.join(backup_dir, f'schedule_backup_{timestamp}.db')

    try:
        shutil.copy2(DB_PATH, backup_path)
        logger.info(f"Database backup created: {backup_path}")

        # Список всех бэкапов
        backups = []
        for f in os.listdir(backup_dir):
            if f.endswith('.db'):
                filepath = os.path.join(backup_dir, f)
                backups.append({
                    'filename': f,
                    'size': os.path.getsize(filepath),
                    'created': datetime.fromtimestamp(os.path.getctime(filepath)).isoformat()
                })

        return jsonify({
            'status': 'success',
            'message': 'Резервная копия создана',
            'backup_path': backup_path,
            'backups': sorted(backups, key=lambda x: x['created'], reverse=True)[:10]
        })

    except Exception as e:
        logger.exception(f"Backup error: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ==========================================
# TUNNEL MONITORING API
# ==========================================

@api_bp.route('/api/tunnel-logs', methods=['GET'])
def api_tunnel_logs():
    """Get tunnel logs"""
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401

    logs = []
    try:
        log_path = os.path.join(LOGS_DIR, 'tunnel.log')
        if os.path.exists(log_path):
            with open(log_path, 'r', encoding='utf-8') as f:
                logs = f.readlines()[-100:]  # Последние 100 строк
    except:
        pass

    return jsonify({
        'status': 'success',
        'logs': logs
    })


@api_bp.route('/api/port-check', methods=['GET'])
def api_port_check():
    """Check if port 8080 is accessible from outside"""
    import socket

    def check_port(host, port, timeout=3):
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(timeout)
            result = sock.connect_ex((host, port))
            sock.close()
            return result == 0
        except:
            return False

    # Проверяем локально
    local_ok = check_port('127.0.0.1', 8080)

    # Проверяем локальный IP
    import subprocess
    try:
        result = subprocess.run(['ipconfig'], capture_output=True, text=True, timeout=5)
        output = result.stdout
        # Ищем IPv4 адрес
        import re
        ip_match = re.search(r'IPv4-адрес.*?(\d+\.\d+\.\d+\.\d+)', output)
        if not ip_match:
            ip_match = re.search(r'IPv4 Address.*?(\d+\.\d+\.\d+\.\d+)', output)
        local_ip = ip_match.group(1) if ip_match else 'unknown'
    except:
        local_ip = 'unknown'

    return jsonify({
        'status': 'success',
        'local_port_8080': 'open' if local_ok else 'closed',
        'local_ip': local_ip,
        'message': 'Порт 8080 открыт локально' if local_ok else 'Порт 8080 закрыт'
    })


# ==========================================
# CHAT VK STATUS API
# ==========================================

@api_bp.route('/api/chat/vk-status', methods=['GET'])
def api_chat_vk_status():
    """Get VK chat connection status"""
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401

    try:
        config_path = os.path.join(DATA_DIR, 'vk_config.json')
        with open(config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)

        return jsonify({
            'status': 'success',
            'vk_connected': bool(config.get('token')),
            'group_id': config.get('group_id'),
            'admin_id': config.get('admin_vk_id'),
            'chat_peer_id': config.get('chat_peer_id')
        })
    except:
        return jsonify({
            'status': 'success',
            'vk_connected': False
        })


# ==========================================
# CONVERTER OPEN API
# ==========================================

@api_bp.route('/api/converter/files', methods=['GET'])
def api_converter_files():
    """Get list of converter files"""
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401
    
    from web_converter import CONVERTER_UPLOADS_DIR
    if not os.path.exists(CONVERTER_UPLOADS_DIR):
        return jsonify({'status': 'success', 'files': []})
    
    files = []
    for f in os.listdir(CONVERTER_UPLOADS_DIR):
        if f.endswith('.xlsx') or f.endswith('.xls'):
            filepath = os.path.join(CONVERTER_UPLOADS_DIR, f)
            files.append({
                'filename': f,
                'size': os.path.getsize(filepath),
                'modified': datetime.fromtimestamp(os.path.getmtime(filepath)).isoformat()
            })
    
    return jsonify({'status': 'success', 'files': files})


@api_bp.route('/api/converter/open', methods=['POST'])
def api_converter_open():
    """Open price file for editing"""
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401

    data = request.json or {}
    filename = data.get('filename')

    if not filename:
        return jsonify({'status': 'error', 'message': 'Filename required'}), 400

    from web_converter import CONVERTER_UPLOADS_DIR
    filepath = os.path.join(CONVERTER_UPLOADS_DIR, filename)

    if not os.path.exists(filepath):
        return jsonify({'status': 'error', 'message': 'File not found'}), 404

    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=False)
        ws = wb.active

        products = []
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row[0]:
                products.append({
                    'row': i,
                    'name': str(row[0]) if row[0] else '',
                    'price': float(row[1]) if len(row) > 1 and row[1] else 0,
                    'unit': str(row[2]) if len(row) > 2 and row[2] else 'шт'
                })

        wb.close()

        return jsonify({
            'status': 'success',
            'products': products,
            'count': len(products)
        })

    except Exception as e:
        logger.exception(f"Converter open error: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ==========================================
# SALARY SUMMARY API — расчёт зарплат
# ==========================================

@api_bp.route('/api/salary-summary', methods=['GET'])
def api_get_salary_summary():
    """Сводка зарплат сотрудников за месяц"""
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401

    if session.get('role') != 'admin':
        return jsonify({'status': 'error', 'message': 'Требуется роль администратора'}), 403

    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)

    if not year or not month:
        now = datetime.now()
        year = year or now.year
        month = month or now.month

    db = get_db_connection()
    cursor = db.cursor()

    # Получаем всех сотрудников
    cursor.execute('SELECT id, full_name, username FROM users WHERE role != "admin" ORDER BY full_name')
    users = cursor.fetchall()

    result = []
    for user in users:
        user_id = user['id']

        # Считаем смены за месяц (используем work_sessions)
        cursor.execute('''
            SELECT COUNT(*) as days, COALESCE(SUM(revenue_total), 0) as total_revenue
            FROM work_sessions
            WHERE user_id = ? AND strftime('%Y', opened_at) = ? AND strftime('%m', opened_at) = ?
        ''', (user_id, str(year), str(month).zfill(2)))
        shift_data = cursor.fetchone()
        days_worked = shift_data['days'] or 0
        total_revenue = shift_data['total_revenue'] or 0

        # Рассчитываем зарплату (10% от выручки)
        salary_due = total_revenue * 0.10

        # Получаем надбавки/штрафы за месяц
        cursor.execute('''
            SELECT COALESCE(SUM(amount), 0) as total
            FROM salary_adjustments
            WHERE user_id = ? AND year = ? AND month = ?
        ''', (user_id, year, month))
        adj_data = cursor.fetchone()
        adjustments = adj_data['total'] or 0

        # Получаем взятые деньги (из work_journal_entries)
        cursor.execute('''
            SELECT COALESCE(SUM(amount), 0) as total
            FROM work_journal_entries
            WHERE shift_id IN (
                SELECT id FROM work_sessions WHERE user_id = ? AND strftime('%Y', opened_at) = ? AND strftime('%m', opened_at) = ?
            ) AND kind = 'salary_taken'
        ''', (user_id, str(year), str(month).zfill(2)))
        taken_data = cursor.fetchone()
        salary_taken = taken_data['total'] or 0

        # Баланс
        balance = salary_due + adjustments - salary_taken

        result.append({
            'user_id': user_id,
            'full_name': user['full_name'] or user['username'],
            'days_worked': days_worked,
            'total_revenue': total_revenue,
            'salary_due': salary_due,
            'adjustments': adjustments,
            'salary_taken': salary_taken,
            'balance': balance
        })

    return jsonify({'status': 'success', 'data': result})


# ==========================================
# EMPLOYEE STATS API — статистика сотрудников
# ==========================================

@api_bp.route('/api/employee-stats', methods=['GET'])
def api_get_employee_stats():
    """Статистика работы сотрудников за месяц"""
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401

    if session.get('role') != 'admin':
        return jsonify({'status': 'error', 'message': 'Требуется роль администратора'}), 403

    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)

    if not year or not month:
        now = datetime.now()
        year = year or now.year
        month = month or now.month

    db = get_db_connection()
    cursor = db.cursor()

    # Получаем всех сотрудников
    cursor.execute('SELECT id, full_name, username FROM users WHERE role != "admin" ORDER BY full_name')
    users = cursor.fetchall()

    result = []
    for user in users:
        user_id = user['id']

        # Считаем смены
        cursor.execute('''
            SELECT
                COUNT(*) as shifts_total,
                COALESCE(SUM(CASE WHEN closed_at IS NOT NULL THEN 1 ELSE 0 END), 0) as shifts_closed,
                COALESCE(SUM(revenue_total), 0) as total_revenue
            FROM work_sessions
            WHERE user_id = ? AND year = ? AND month = ?
        ''', (user_id, year, month))
        shift_data = cursor.fetchone()
        shifts_total = shift_data['shifts_total'] or 0
        shifts_closed = shift_data['shifts_closed'] or 0
        total_revenue = shift_data['total_revenue'] or 0
        avg_revenue = total_revenue / shifts_total if shifts_total > 0 else 0

        # Опоздания (открытие после 9:00)
        cursor.execute('''
            SELECT COUNT(*) as cnt
            FROM work_sessions
            WHERE user_id = ? AND year = ? AND month = ?
            AND time(opened_at) > '09:00:00'
        ''', (user_id, year, month))
        late_openings = cursor.fetchone()['cnt'] or 0

        # Ранние закрытия (до 19:00)
        cursor.execute('''
            SELECT COUNT(*) as cnt
            FROM work_sessions
            WHERE user_id = ? AND year = ? AND month = ?
            AND closed_at IS NOT NULL
            AND time(closed_at) < '19:00:00'
        ''', (user_id, year, month))
        early_closings = cursor.fetchone()['cnt'] or 0

        # Поздние закрытия (после 19:00)
        cursor.execute('''
            SELECT COUNT(*) as cnt
            FROM work_sessions
            WHERE user_id = ? AND year = ? AND month = ?
            AND closed_at IS NOT NULL
            AND time(closed_at) > '19:00:00'
        ''', (user_id, year, month))
        late_closings = cursor.fetchone()['cnt'] or 0

        # Расхождения по кассе
        cursor.execute('''
            SELECT COUNT(*) as cnt
            FROM work_journal_entries
            WHERE shift_id IN (
                SELECT id FROM work_sessions WHERE user_id = ? AND year = ? AND month = ?
            ) AND kind = 'discrepancy'
        ''', (user_id, year, month))
        cash_discrepancies = cursor.fetchone()['cnt'] or 0

        result.append({
            'user_id': user_id,
            'full_name': user['full_name'] or user['username'],
            'shifts_total': shifts_total,
            'shifts_closed': shifts_closed,
            'avg_revenue': avg_revenue,
            'total_revenue': total_revenue,
            'late_openings': late_openings,
            'early_closings': early_closings,
            'late_closings': late_closings,
            'cash_discrepancies': cash_discrepancies
        })

    return jsonify({'status': 'success', 'data': result})


# ==========================================
# SYSTEM UPDATE API — загрузка обновлений
# ==========================================

@api_bp.route('/api/system/update', methods=['POST'])
def api_system_update():
    """Загрузка файла обновления системы"""
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401

    if session.get('role') != 'admin':
        return jsonify({'status': 'error', 'message': 'Требуется роль администратора'}), 403

    if 'file' not in request.files:
        return jsonify({'status': 'error', 'message': 'Файл не найден'}), 400

    file = request.files['file']
    if not file.filename.endswith('.exe'):
        return jsonify({'status': 'error', 'message': 'Требуется файл .exe'}), 400

    import shutil
    from datetime import datetime

    # Папка для обновлений
    update_dir = os.path.join(DATA_DIR, 'updates')
    os.makedirs(update_dir, exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'update_{timestamp}.exe'
    filepath = os.path.join(update_dir, filename)

    try:
        file.save(filepath)
        logger.info(f"Update file uploaded: {filepath}")

        # Запись в audit_log
        audit_log(session['user_id'], 'system_update', {
            'action': 'upload',
            'filename': filename,
            'filepath': filepath
        })

        # TODO: Добавить автоматическую установку обновления
        # Для этого нужно:
        # 1. Остановить сервер
        # 2. Заменить exe файл
        # 3. Перезапустить сервер

        return jsonify({
            'status': 'success',
            'message': 'Обновление загружено',
            'filename': filename,
            'filepath': filepath
        })

    except Exception as e:
        logger.exception(f"Update upload error: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ==========================================
# VK MEMBERS API — участники ВКонтакте
# ==========================================

@api_bp.route('/api/vk-members', methods=['GET'])
def api_get_vk_members():
    """Получить участников чата ВКонтакте"""
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not authorized'}), 401

    if session.get('role') != 'admin':
        return jsonify({'status': 'error', 'message': 'Требуется роль администратора'}), 403

    import json

    # Читаем VK конфиг
    vk_config_path = os.path.join(DATA_DIR, 'vk_config.json')
    try:
        with open(vk_config_path, 'r', encoding='utf-8') as f:
            vk_config = json.load(f)
    except:
        return jsonify({'status': 'error', 'message': 'VK конфиг не найден'}), 404

    group_id = vk_config.get('group_id')
    token = vk_config.get('token')

    if not group_id or not token:
        return jsonify({'status': 'error', 'message': 'VK токен или group_id не настроены'}), 400

    # Получаем участников беседы через VK API
    import urllib.request
    try:
        # Метод messages.getConversationMembers
        url = f'https://api.vk.com/method/messages.getConversationMembers?peer_id=2000000001&v=5.131&access_token={token}'
        req = urllib.request.Request(url)
        with urllib.request.urlopen(req, timeout=10) as response:
            data = json.loads(response.read().decode('utf-8'))

        if 'error' in data:
            return jsonify({'status': 'error', 'message': data['error'].get('error_msg', 'VK API error')}), 400

        items = data.get('response', {}).get('items', [])
        members = []
        for item in items:
            member_id = item.get('member_id')
            # member_id > 0 — пользователь, < 0 — бот/чат
            if member_id and member_id > 0:
                members.append({
                    'vk_id': member_id,
                    'name': item.get('name', 'Unknown')
                })

        return jsonify({'status': 'success', 'members': members, 'count': len(members)})

    except Exception as e:
        logger.exception(f"VK members fetch error: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500
